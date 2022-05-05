#    TBXTools
#    version: 2022/05/05
#    Copyright: Antoni Oliver (2022) - Universitat Oberta de Catalunya - aoliverg@uoc.edu
#    This program is free software: you can redistribute it and/or modify
#    it under the terms of the GNU General Public License as published by
#    the Free Software Foundation, either version 3 of the License, or
#    (at your option) any later version.

#    This program is distributed in the hope that it will be useful,
#    but WITHOUT ANY WARRANTY; without even the implied warranty of
#    MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
#    GNU General Public License for more details.

#    You should have received a copy of the GNU General Public License
#    along with this program.  If not, see <https://www.gnu.org/licenses/>.

import os
import codecs
import sqlite3
import xml.etree.cElementTree as etree

import nltk
from nltk.util import ngrams
from nltk.probability import FreqDist
from nltk.collocations import *
import re
import pickle
import gzip
import operator
import sys
import math
import csv

import string

import importlib

import gensim
from gensim.models import Word2Vec
from gensim.models import KeyedVectors
import numpy
import collections
import numpy as np

import time

try:
    import spacy
except:
    pass
    
try:
    import spacy_udpipe
except:
    pass
import subprocess
import openpyxl
from openpyxl import load_workbook

class TBXTools:
    '''Class for automatic terminology extraction and terminology management.'''
    def __init__(self):
        self.maxinserts=10000 #controls the maximum number of inserts in memory
        self.sl_lang=""
        self.tl_lang=""
        self.max_id_corpus=0
        
        self.sl_stopwords=[]
        self.tl_stopwords=[]
        self.sl_inner_stopwords=[]
        self.tl_inner_stopwords=[]
        self.sl_exclsions_regexps=[]
        self.tl_exclusion_regexps=[]
        self.sl_morphonorm_rules=[]
        self.tl_morphonorm_rules=[]
        self.evaluation_terms={}
        self.tsr_terms=[]
        self.exclusion_terms={}
        self.exclusion_no_terms={}
        self.ngrams={}
        self.tagged_ngrams={}
        self.term_candidates={}
        self.linguistic_patterns={}
        
        self.knownterms=[]
        self.n_min=1
        self.n_max=5
        
        self.n_min_pos_patterns=1000
        self.n_max_pos_patterns=1
        
        self.punctuation=string.punctuation
        self.sl_stopwords.extend(self.punctuation)
        self.tl_stopwords.extend(self.punctuation)
        self.sl_inner_stopwords.extend(self.punctuation)
        self.tl_inner_stopwords.extend(self.punctuation)
        
        self.specificSLtokenizer=False
        self.specificTLtokenizer=False
        
        self.SLtokenizer=None
        self.TLtokenizer=None
        
        
        
        
    def create_project(self,project_name,sllang=None, tllang=None,overwrite=False):
        '''Opens a project. If the project already exists, it raises an exception. To avoid the exception use overwrite=True. To open existing projects, use the open_project method.'''
        #sllang and tllang are not longer used.
        if os.path.isfile(project_name) and not overwrite:
                raise Exception("This file already exists")
        
        else:
            if os.path.isfile(project_name) and overwrite:
                os.remove(project_name)
            self.conn=sqlite3.connect(project_name)
            self.cur = self.conn.cursor() 
            self.cur2 = self.conn.cursor()
            with self.conn:
                self.cur = self.conn.cursor()
                self.cur.execute("CREATE TABLE sl_corpus(id INTEGER PRIMARY KEY AUTOINCREMENT, segment TEXT)")
                self.cur.execute("CREATE TABLE tl_corpus(id INTEGER PRIMARY KEY AUTOINCREMENT, segment TEXT)")
                self.cur.execute("CREATE TABLE parallel_corpus(id INTEGER PRIMARY KEY AUTOINCREMENT, segmentSL, segmentTL TEXT)")
                self.cur.execute("CREATE TABLE tagged_parallel_corpus(id INTEGER PRIMARY KEY, tagged_segmentSL, tagged_segmentTL TEXT)")
                self.cur.execute("CREATE TABLE sl_corpus_c(id INTEGER PRIMARY KEY AUTOINCREMENT, segment TEXT)")
                self.cur.execute("CREATE TABLE tl_corpus_c(id INTEGER PRIMARY KEY AUTOINCREMENT, segment TEXT)")
                self.cur.execute("CREATE TABLE sl_tagged_corpus(id INTEGER PRIMARY KEY AUTOINCREMENT, tagged_segment TEXT)")
                self.cur.execute("CREATE TABLE tl_tagged_corpus(id INTEGER PRIMARY KEY AUTOINCREMENT, tagged_segment TEXT)")
                self.cur.execute("CREATE TABLE sl_tagged_corpus_c(id INTEGER PRIMARY KEY AUTOINCREMENT, tagged_segment TEXT)")
                self.cur.execute("CREATE TABLE tl_tagged_corpus_c(id INTEGER PRIMARY KEY AUTOINCREMENT, tagged_segment TEXT)")
                self.cur.execute("CREATE TABLE sl_stopwords (id INTEGER PRIMARY KEY AUTOINCREMENT, sl_stopword TEXT)")
                self.cur.execute("CREATE TABLE sl_inner_stopwords (id INTEGER PRIMARY KEY AUTOINCREMENT, sl_inner_stopword TEXT)")
                self.cur.execute("CREATE TABLE tl_stopwords (id INTEGER PRIMARY KEY AUTOINCREMENT, tl_stopword TEXT)")
                self.cur.execute("CREATE TABLE tl_inner_stopwords (id INTEGER PRIMARY KEY AUTOINCREMENT, tl_inner_stopword TEXT)")
                self.cur.execute("CREATE TABLE sl_exclusion_regexps (id INTEGER PRIMARY KEY AUTOINCREMENT, sl_exclusion_regexp TEXT)")
                self.cur.execute("CREATE TABLE tl_exclusion_regexps (id INTEGER PRIMARY KEY AUTOINCREMENT, tl_exclusion_regexp TEXT)")
                self.cur.execute("CREATE TABLE sl_morphonorm_rules (id INTEGER PRIMARY KEY AUTOINCREMENT, sl_morphonorm_rule TEXT)")
                self.cur.execute("CREATE TABLE tl_morphonorm_rules (id INTEGER PRIMARY KEY AUTOINCREMENT, tl_morphonorm_rule TEXT)")
                self.cur.execute("CREATE TABLE evaluation_terms (id INTEGER PRIMARY KEY AUTOINCREMENT, sl_term TEXT, tl_term TEXT)")
                self.cur.execute("CREATE TABLE reference_terms (id INTEGER PRIMARY KEY AUTOINCREMENT, sl_term TEXT, tl_term TEXT)")
                self.cur.execute("CREATE TABLE validated_terms (id INTEGER PRIMARY KEY AUTOINCREMENT, sl_term TEXT, tl_term TEXT)")
                self.cur.execute("CREATE TABLE compoundify_terms_sl (id INTEGER PRIMARY KEY AUTOINCREMENT, term TEXT)")
                self.cur.execute("CREATE TABLE compoundify_terms_tl (id INTEGER PRIMARY KEY AUTOINCREMENT, term TEXT)")
                self.cur.execute("CREATE TABLE tsr_terms (id INTEGER PRIMARY KEY AUTOINCREMENT, term TEXT)")
                self.cur.execute("CREATE TABLE tosearch_terms (id INTEGER PRIMARY KEY AUTOINCREMENT, term TEXT)")
                self.cur.execute("CREATE TABLE exclusion_terms (id INTEGER PRIMARY KEY AUTOINCREMENT, sl_term TEXT, tl_term TEXT)")
                self.cur.execute("CREATE TABLE exclusion_noterms (id INTEGER PRIMARY KEY AUTOINCREMENT, sl_term TEXT, tl_term TEXT)")
                self.cur.execute("CREATE TABLE tokens (id INTEGER PRIMARY KEY AUTOINCREMENT, token TEXT, frequency INTEGER)")
                self.cur.execute("CREATE TABLE ngrams (id INTEGER PRIMARY KEY AUTOINCREMENT, ngram TEXT, n INTEGER, frequency INTEGER)")
                self.cur.execute("CREATE TABLE tagged_ngrams (id INTEGER PRIMARY KEY AUTOINCREMENT, ngram TEXT, tagged_ngram TEXT, n INTEGER, frequency INTEGER)")
                
                self.cur.execute("CREATE INDEX indextaggedngram on tagged_ngrams (ngram);")
                
                self.cur.execute("CREATE TABLE embeddings_sl (id INTEGER PRIMARY KEY AUTOINCREMENT, candidate TEXT, embedding BLOB)")
                self.cur.execute("CREATE INDEX indexembeddings_sl on embeddings_sl (candidate);")
                
                self.cur.execute("CREATE TABLE embeddings_sl_ref (id INTEGER PRIMARY KEY AUTOINCREMENT, candidate TEXT, embedding BLOB)")
                self.cur.execute("CREATE INDEX indexembeddings_sl_ref on embeddings_sl_ref (candidate);")
                
                self.cur.execute("CREATE TABLE embeddings_tl (id INTEGER PRIMARY KEY AUTOINCREMENT, candidate TEXT, embedding BLOB)")
                self.cur.execute("CREATE INDEX indexembeddings_tl on embeddings_tl (candidate);")
                
                self.cur.execute("CREATE TABLE term_candidates (id INTEGER PRIMARY KEY AUTOINCREMENT, candidate TEXT, n INTEGER, frequency INTEGER, measure TEXT, value FLOAT)")
                self.cur.execute("CREATE TABLE index_pt(id INTEGER PRIMARY KEY AUTOINCREMENT, source TEXT, target TEXT, probability FLOAT)")
                self.cur.execute("CREATE INDEX index_index_pt on index_pt (source);")
                self.cur.execute("CREATE TABLE linguistic_patterns (id INTEGER PRIMARY KEY AUTOINCREMENT, linguistic_pattern TEXT)")
                
                self.conn.commit()
                
    def open_project(self,project_name):
        '''Opens an existing project. If the project doesn't exist it raises an exception.'''
        if not os.path.isfile(project_name):
                raise Exception("Project not found")
        else:
            self.conn=sqlite3.connect(project_name)
            self.cur = self.conn.cursor() 
            self.cur2 = self.conn.cursor()
            

    #METODES DELETES
    def delete_configuration(self):
        '''Deletes the project configuration.'''
        with self.conn:
            self.cur.execute('DELETE FROM configuration')
            self.conn.commit()
    
    def delete_sl_corpus(self):
        '''Deletes de source language corpus.'''
        with self.conn:
            self.cur.execute('DELETE FROM sl_corpus')
            self.conn.commit()
    
    def delete_tl_corpus(self):
        '''Deletes de target language corpus.'''
        with self.conn:
            self.cur.execute('DELETE FROM tl_corpus')
            self.conn.commit()
            
    def delete_parallel_corpus(self):
        '''Deletes de target language corpus.'''
        with self.conn:
            self.cur.execute('DELETE FROM parallel_corpus')
            self.conn.commit()
            
    def delete_sl_corpus_c(self):
        '''Deletes de source language contrast corpus.'''
        with self.conn:
            self.cur.execute('DELETE FROM sl_corpus_c')
            self.conn.commit()
    
    def delete_tl_corpus_c(self):
        '''Deletes de target language contrast corpus.'''
        with self.conn:
            self.cur.execute('DELETE FROM tl_corpus_c')
            self.conn.commit()
    
    def delete_sl_tagged_corpus(self):
        '''Deletes the source language tagged corpus.'''
        with self.conn:
            self.cur.execute('DELETE FROM sl_tagged_corpus')
            self.conn.commit()
    
    def delete_tl_tagged_corpus(self):
        '''Deletes the target language tagged corpus.'''
        with self.conn:
            self.cur.execute('DELETE FROM tl_tagged_corpus')
            self.conn.commit()
         
    def delete_sl_tagged_corpus_c(self):
        '''Deletes the source language contrast tagged corpus.'''
        with self.conn:
            self.cur.execute('DELETE FROM sl_tagged_corpus_c')
            self.conn.commit()
    
    def delete_tl_tagged_corpus_c(self):
        '''Deletes the target language contrast tagged corpus.'''
        with self.conn:
            self.cur.execute('DELETE FROM tl_tagged_corpus_c')
            self.conn.commit()
    
    def delete_sl_stopwords(self):
        '''Deletes the stop-words for the source language.'''
        #self.sl_stopwords=[]
        with self.conn:
            self.cur.execute('DELETE FROM sl_stopwords')
            self.conn.commit()
            
    def delete_tl_stopwords(self):
        '''Deletes the stop-words fot the target language.'''
        #self.tl_stopwords=[]
        with self.conn:
            self.cur.execute('DELETE FROM tl_stopwords')
            self.conn.commit()
            
    def delete_sl_inner_stopwords(self):
        '''Deletes the inner stop-words for the source language.'''
        #self.sl_inner_stopwords=[]
        with self.conn:
            self.cur.execute('DELETE FROM sl_inner_stopwords')
            self.conn.commit()
            
    def delete_tl_inner_stopwords(self):
        '''Deletes the innter stop-words for the target language.'''
        #self.tl_inner_stopwords=[]
        with self.conn:
            self.cur.execute('DELETE FROM tl_inner_stopwords')
            self.conn.commit()
    
    def delete_sl_exclusion_regexps(self):
        '''Deletes the exclusion regular expressions for the source language.'''
        #self.sl_exclusion_regexps=[]
        with self.conn:
            self.cur.execute('DELETE FROM sl_exclusion_regexps')
            self.conn.commit()
    
    def delete_tl_exclusion_regexps(self):
        '''Deletes the exclusion regular expressions for the target language.'''
        #self.tl_exclusion_regexps=[]
        with self.conn:
            self.cur.execute('DELETE FROM tl_exclusion_regexps')
            self.conn.commit()
            
    def delete_sl_morphonorm_rules(self):
        '''Deletes the morphological normalisation rules for the source language.'''
        #self.sl_morphonorm_rules=[]
        with self.conn:
            self.cur.execute('DELETE FROM sl_morphonorm_rules')
            self.conn.commit()
            
    def delete_tl_morphonorm_rules(self):
        '''Deletes the morphological normalisation rules for the target language.'''
        #self.tl_morphonorm_rules=[]
        with self.conn:
            self.cur.execute('DELETE FROM tl_morphonorm_rules')
            self.conn.commit()

    def delete_evaluation_terms(self):
        '''Deletes the evaluation terms.'''
        #self.evaluation_terms={}
        with self.conn:
            self.cur.execute('DELETE FROM evaluation_terms')
            self.conn.commit()
    
    def delete_reference_terms(self):
        '''Deletes the reference terms.'''
        #self.evaluation_terms={}
        with self.conn:
            self.cur.execute('DELETE FROM reference_terms')
            self.conn.commit()
            
    def delete_validated_terms(self):
        '''Deletes the validated terms.'''
        #self.evaluation_terms={}
        with self.conn:
            self.cur.execute('DELETE FROM validated_terms')
            self.conn.commit()
            
    def delete_compoundify_terms_sl(self):
        '''Deletes the compoundify terms for the source language.'''
        #self.exclusion_terms={}
        with self.conn:
            self.cur.execute('DELETE FROM compoundify_terms_sl')
            self.conn.commit() 
    
    def delete_compoundify_terms_tl(self):
        '''Deletes the compoundify terms for the target language.'''
        #self.exclusion_terms={}
        with self.conn:
            self.cur.execute('DELETE FROM compoundify_terms_sl')
            self.conn.commit()
    
    def delete_tsr_terms(self):
        '''Deletes the TSR terms.'''
        #self.exclusion_terms={}
        with self.conn:
            self.cur.execute('DELETE FROM tsr_terms')
            self.conn.commit()
    
    def delete_exclusion_terms(self):
        '''Deletes the exclusion terms.'''
        #self.exclusion_terms={}
        with self.conn:
            self.cur.execute('DELETE FROM exclusion_terms')
            self.conn.commit()
    
    def delete_exclusion_no_terms(self):
        '''Deletes the exclusion no terms.'''
        #self.exclusion_terms={}
        with self.conn:
            self.cur.execute('DELETE FROM exclusion_no_terms')
            self.conn.commit() 
            
    def delete_tokens(self):
        '''Deletes the tokens.'''
        #self.ngrams={}
        with self.conn:
            self.cur.execute('DELETE FROM tokens')
            self.conn.commit()
            
    def delete_ngrams(self):
        '''Deletes the ngrams.'''
        #self.ngrams={}
        with self.conn:
            self.cur.execute('DELETE FROM ngrams')
            self.conn.commit()
            
    def delete_tagged_ngrams(self):
        '''Deletes the tagged ngrams.'''
        #self.tagged_ngrams={}
        with self.conn:
            self.cur.execute('DELETE FROM tagged_ngrams')
            self.conn.commit()
            
    def delete_embeddings_sl(self):
        '''Deletes the embeddings for the source language.'''
        #self.tagged_ngrams={}
        with self.conn:
            self.cur.execute('DELETE FROM embeddings_sl')
            self.conn.commit()
            
    def delete_embeddings_tl(self):
        '''Deletes the embeddings for the target language.'''
        #self.tagged_ngrams={}
        with self.conn:
            self.cur.execute('DELETE FROM embeddings_tl')
            self.conn.commit()
            
    def delete_term_candidates(self):
        '''Deletes the term candidates.'''
        #self.term_candidates={}
        with self.conn:
            self.cur.execute('DELETE FROM term_candidates')
            self.conn.commit()
            
    def delete_linguistic_patterns(self):
        '''Deletes the linguistic patterns for linguistic terminology extraction.'''
        #self.exclusion_terms={}
        with self.conn:
            self.cur.execute('DELETE FROM linguistic_patterns')
            self.conn.commit() 
                     
    def load_sl_corpus(self,corpusfile, encoding="utf-8", compoundify=False, comp_symbol="▁"):
        '''Loads a monolingual corpus for the source language. It's recommended, but not compulsory, that the corpus is segmented (one segment per line). Use external tools to segment the corpus. A plain text corpus (not segmented), can be aslo used.'''
        if compoundify:
            compterms=[]
            self.cur.execute('SELECT term from compoundify_terms_sl')
            data=self.cur.fetchall()
            for d in data:
                compterms.append(d[0])            
        cf=codecs.open(corpusfile,"r",encoding=encoding,errors="ignore")
        data=[]
        continserts=0
        for line in cf:            
            record=[]
            line=line.rstrip()
            if compoundify:
                for compterm in compterms:
                    if line.find(compterm)>=1:
                        comptermMOD=compterm.replace(" ",comp_symbol)
                        line=line.replace(compterm,comptermMOD)
            record.append(line)
            data.append(record)
            continserts+=1
            if continserts==self.maxinserts:
                self.cur.executemany("INSERT INTO sl_corpus (segment) VALUES (?)",data)
                data=[]
                continserts=0
        with self.conn:
            self.cur.executemany("INSERT INTO sl_corpus (segment) VALUES (?)",data)    
        self.conn.commit()
        
    def load_tl_corpus(self,corpusfile, encoding="utf-8", compoundify=False, comp_symbol="▁"):
        '''Loads a monolingual corpus for the target language. It's recommended, but not compulsory, that the corpus is segmented (one segment per line). Use TBXTools external tools to segment the corpus. A plain text corpus (not segmented), can be aslo used.'''
        
        if compoundify:
            compterms=[]
            self.cur.execute('SELECT term from compoundify_terms_tl')
            data=self.cur.fetchall()
            for d in data:
                compterms.append(d[0])            
        cf=codecs.open(corpusfile,"r",encoding=encoding,errors="ignore")
        data=[]
        continserts=0
        for line in cf:            
            record=[]
            line=line.rstrip()
            if compoundify:
                for compterm in compterms:
                    if line.find(compterm)>=1:
                        comptermMOD=compterm.replace(" ",comp_symbol)
                        line=line.replace(compterm,comptermMOD)
            record.append(line)
            data.append(record)
            continserts+=1
            if continserts==self.maxinserts:
                self.cur.executemany("INSERT INTO tl_corpus (segment) VALUES (?)",data)
                data=[]
                continserts=0
        with self.conn:
            self.cur.executemany("INSERT INTO tl_corpus (segment) VALUES (?)",data)    
        self.conn.commit()        
    
    def load_sl_corpus_c(self,corpusfile, encoding="utf-8", compoundify=False, comp_symbol="▁"):
        '''Loads a monolingual contrast corpus for the source language. It's recommended, but not compulsory, that the corpus is segmented (one segment per line). Use external tools to segment the corpus. A plain text corpus (not segmented), can be aslo used.'''
        if compoundify:
            compterms=[]
            self.cur.execute('SELECT term from compoundify_terms_sl')
            data=self.cur.fetchall()
            for d in data:
                compterms.append(d[0])            
        cf=codecs.open(corpusfile,"r",encoding=encoding,errors="ignore")
        data=[]
        continserts=0
        for line in cf:            
            record=[]
            line=line.rstrip()
            if compoundify:
                for compterm in compterms:
                    if line.find(compterm)>=1:
                        comptermMOD=compterm.replace(" ",comp_symbol)
                        line=line.replace(compterm,comptermMOD)
            record.append(line)
            data.append(record)
            continserts+=1
            if continserts==self.maxinserts:
                cur.executemany("INSERT INTO sl_corpus_c (segment) VALUES (?)",data)
                data=[]
                continserts=0
        with self.conn:
            self.cur.executemany("INSERT INTO sl_corpus_c (segment) VALUES (?)",data)    
        self.conn.commit()
        
    def load_tl_corpus_c(self,corpusfile, encoding="utf-8", compoundify=False, comp_symbol="▁"):
        '''Loads a monolingual contrast corpus for the target language. It's recommended, but not compulsory, that the corpus is segmented (one segment per line). Use TBXTools external tools to segment the corpus. A plain text corpus (not segmented), can be aslo used.'''
        
        if compoundify:
            compterms=[]
            self.cur.execute('SELECT term from compoundify_terms_tl')
            data=self.cur.fetchall()
            for d in data:
                compterms.append(d[0])            
        cf=codecs.open(corpusfile,"r",encoding=encoding,errors="ignore")
        data=[]
        continserts=0
        for line in cf:            
            record=[]
            line=line.rstrip()
            if compoundify:
                for compterm in compterms:
                    if line.find(compterm)>=1:
                        comptermMOD=compterm.replace(" ",comp_symbol)
                        line=line.replace(compterm,comptermMOD)
            record.append(line)
            data.append(record)
            continserts+=1
            if continserts==self.maxinserts:
                cur.executemany("INSERT INTO tl_corpus_c (segment) VALUES (?)",data)
                data=[]
                continserts=0
        with self.conn:
            self.cur.executemany("INSERT INTO tl_corpus_c (segment) VALUES (?)",data)    
        self.conn.commit() 
        
    def load_parallel_corpus_Moses(self,slcorpusfile, tlcorpusfile, feed_monolingual=True, encoding="utf-8"):
        '''Loads a parallel corpus in Moses format (that is, in two independent files). It expects one segment per line.'''
        slcf=codecs.open(slcorpusfile,"r",encoding=encoding)
        tlcf=codecs.open(tlcorpusfile,"r",encoding=encoding)
        parallel_data=[]
        sl_data=[]
        tl_data=[]
        parallel_data=[]
        continserts=0
        while 1:
            sl_segment=slcf.readline()
            if not sl_segment:
                break
            tl_segment=tlcf.readline()
            continserts+=1            
            sl_record=[]
            tl_record=[]
            parallel_record=[]
            sl_segment=sl_segment.rstrip()
            tl_segment=tl_segment.rstrip()
            parallel_record.append(sl_segment)
            parallel_record.append(tl_segment)
            sl_record.append(sl_segment)
            tl_record.append(tl_segment)
            parallel_data.append(parallel_record)
            sl_data.append(sl_record)
            tl_data.append(tl_record)
            if continserts==self.maxinserts:
                self.cur.executemany("INSERT INTO parallel_corpus (segmentSL, segmentTL) VALUES (?,?)",parallel_data)
                if feed_monolingual:
                    self.cur.executemany("INSERT INTO sl_corpus (segment) VALUES (?)",sl_data)
                    self.cur.executemany("INSERT INTO tl_corpus (segment) VALUES (?)",tl_data)
                parallel_data=[]    
                sl_data=[]
                tl_data=[]
                continserts=0
        with self.conn:
            self.cur.executemany("INSERT INTO parallel_corpus (segmentSL, segmentTL) VALUES (?,?)",parallel_data)
            if feed_monolingual:
                self.cur.executemany("INSERT INTO sl_corpus (segment) VALUES (?)",sl_data)
                self.cur.executemany("INSERT INTO tl_corpus (segment) VALUES (?)",tl_data) 
        self.conn.commit()
        
    def load_parallel_corpus_tabtxt(self,corpusfile, feed_monolingual=True, reverse=False, encoding="utf-8"):
        '''Loads a parallel corpus in tabbed text format (that is, in two independent files). It expects one segment per line.'''
        cf=codecs.open(corpusfile,"r",encoding=encoding)        
        parallel_data=[]
        sl_data=[]
        tl_data=[]
        parallel_data=[]
        parallel_data_rev=[]
        continserts=0
        for linia in cf:
            linia=linia.rstrip()
            camps=linia.split("\t")
            if len(camps)>=2:
                sl_segment=camps[0]
                tl_segment=camps[1]
                continserts+=1            
                sl_record=[]
                tl_record=[]
                parallel_record=[]
                parallel_record_rev=[]
                sl_segment=sl_segment.rstrip()
                tl_segment=tl_segment.rstrip()
                parallel_record.append(sl_segment)
                parallel_record.append(tl_segment)
                parallel_record_rev.append(tl_segment)
                parallel_record_rev.append(sl_segment)
                sl_record.append(sl_segment)
                tl_record.append(tl_segment)
                parallel_data.append(parallel_record)
                parallel_data_rev.append(parallel_record_rev)
                sl_data.append(sl_record)
                tl_data.append(tl_record)
            if continserts==self.maxinserts:
                if reverse:
                    self.cur.executemany("INSERT INTO parallel_corpus (segmentSL, segmentTL) VALUES (?,?)",parallel_data_rev)
                else:
                    self.cur.executemany("INSERT INTO parallel_corpus (segmentSL, segmentTL) VALUES (?,?)",parallel_data)
                if feed_monolingual:
                    if reverse:
                        self.cur.executemany("INSERT INTO sl_corpus (segment) VALUES (?)",tl_data)
                        self.cur.executemany("INSERT INTO tl_corpus (segment) VALUES (?)",sl_data)
                    else:
                        self.cur.executemany("INSERT INTO sl_corpus (segment) VALUES (?)",sl_data)
                        self.cur.executemany("INSERT INTO tl_corpus (segment) VALUES (?)",tl_data)
                parallel_data=[]
                parallel_data_rev=[]                
                sl_data=[]
                tl_data=[]
                continserts=0
        with self.conn:
            if reverse:
                self.cur.executemany("INSERT INTO parallel_corpus (segmentSL, segmentTL) VALUES (?,?)",parallel_data_rev)
            else:
                self.cur.executemany("INSERT INTO parallel_corpus (segmentSL, segmentTL) VALUES (?,?)",parallel_data)
            if feed_monolingual:
                if reverse:
                    self.cur.executemany("INSERT INTO sl_corpus (segment) VALUES (?)",tl_data)
                    self.cur.executemany("INSERT INTO tl_corpus (segment) VALUES (?)",sl_data)
                else:
                    self.cur.executemany("INSERT INTO sl_corpus (segment) VALUES (?)",sl_data)
                    self.cur.executemany("INSERT INTO tl_corpus (segment) VALUES (?)",tl_data) 
        self.conn.commit()
                    
    def load_parallel_corpus_tmx(self,tmx_file, sl_code="", tl_code="", feed_monolingual=True):
        '''Loads a parallel corpus from a TMX file. Source and target language codes should be given. The codes must be the exactly the same as in the TMX file. A list of codes separated by comma is allowed. '''
        continserts=0
        slcodes=[]
        for slc in sl_code.split(","):
            slcodes.append(slc.strip())
        tlcodes=[]
        for tlc in tl_code.split(","):
            tlcodes.append(tlc.strip())
        data1=[]
        data2=[]
        datap=[]
        sl_segment=""
        tl_segment=""
        current_lang=""
        for event, elem in etree.iterparse(tmx_file,events=("start","end")):
            if event=='start':
                if elem.tag=="tu" and not sl_segment=="" and not tl_segment=="":
                    continserts+=1
                    
                    record1=[]
                    record2=[]
                    recordp=[]
                    record1.append(sl_segment)
                    data1.append(record1)
                    record2.append(tl_segment)
                    data2.append(record2)
                    recordp.append(sl_segment)
                    recordp.append(tl_segment)
                    datap.append(recordp)
                    sl_segment=""
                    tl_segment=""
                    if continserts==self.maxinserts:
                        self.cur.executemany("INSERT INTO parallel_corpus (segmentSL, segmentTL) VALUES (?,?)",datap)
                        if feed_monolingual:
                            self.cur.executemany("INSERT INTO sl_corpus (segment) VALUES (?)",data1) 
                            self.cur.executemany("INSERT INTO tl_corpus (segment) VALUES (?)",data2)
                        data1=[]
                        data2=[]
                        datap=[]
                        continserts=0
                        self.conn.commit()
                elif elem.tag=="tuv":
                    current_lang=elem.attrib['{http://www.w3.org/XML/1998/namespace}lang']
                elif elem.tag=="seg":
                    if elem.text:
                        segmentext=elem.text
                    else:
                        segmentext=""
                    if current_lang in slcodes:
                        sl_segment=segmentext
                    if current_lang in tlcodes:
                        tl_segment=segmentext 
        with self.conn:
            self.cur.executemany("INSERT INTO parallel_corpus (segmentSL, segmentTL) VALUES (?,?)",datap)
            if feed_monolingual:
                self.cur.executemany("INSERT INTO sl_corpus (segment) VALUES (?)",data1)
                self.cur.executemany("INSERT INTO tl_corpus (segment) VALUES (?)",data2) 
        self.conn.commit()                
                        
    def load_parallel_corpus_sdltm(self,sdltmfile, feed_monolingual=True):
        '''Loads a parallel corpus from a SDLTM file.'''
        
        connSDLTM=sqlite3.connect(sdltmfile)
        curSDLTM = connSDLTM.cursor() 
        curSDLTM.execute('select source_segment,target_segment from translation_units;')
        dataSDLTM=curSDLTM.fetchall()
        data1=[]
        data2=[]
        datap=[]
        continserts=0
        for d in dataSDLTM:
            ssxml=d[0]
            tsxml=d[1]
            record1=[]
            record2=[]
            recordp=[]
            try:
                rootSL = etree.fromstring(ssxml)
                for text in rootSL.iter('Value'):
                    sltext="".join(text.itertext()).replace("\n"," ")
                rootTL = etree.fromstring(tsxml)
                for text in rootTL.iter('Value'):
                    tltext="".join(text.itertext()).replace("\n"," ")
                if not sltext=="" and not tltext=="":
                    continserts+=1
                    record1.append(sltext)
                    data1.append(record1)
                    record2.append(tltext)
                    data2.append(record2)
                    recordp.append(sltext)
                    recordp.append(tltext)    
                    datap.append(recordp)
            except:
                print("ERROR")
            if continserts==self.maxinserts:
                self.cur.executemany("INSERT INTO parallel_corpus (segmentSL, segmentTL) VALUES (?,?)",datap)
                if feed_monolingual:
                    self.cur.executemany("INSERT INTO sl_corpus (segment) VALUES (?)",data1) 
                    self.cur.executemany("INSERT INTO tl_corpus (segment) VALUES (?)",data2)
                data1=[]
                data2=[]
                datap=[]
                continserts=0
                self.conn.commit()
        with self.conn:
            self.cur.executemany("INSERT INTO parallel_corpus (segmentSL, segmentTL) VALUES (?,?)",datap)
            if feed_monolingual:
                self.cur.executemany("INSERT INTO sl_corpus (segment) VALUES (?)",data1)
                self.cur.executemany("INSERT INTO tl_corpus (segment) VALUES (?)",data2) 
            self.conn.commit()  
        
        
                        
    def load_sl_tagged_corpus(self,corpusfile,format="TBXTools",encoding="utf-8"):
        '''Loads the source language tagged corpus. 3 formats are allowed:
        - TBXTools: The internal format used by TBXTools. One tagged segment per line.
                f1|l1|t1|p1 f2|l2|t2|p2 ... fn|ln|tn|pn 
        - Freeling: One token per line and segments separated by blank lines
                f1 l1 t1 p1
                f2 l2 t2 p2
                ...
                fn ln tn pn
        - Conll: One of the output formats guiven by the Standford Core NLP analyzer.  On token per line and segments separated by blank lines
                id1 f1 l1 t1 ...
                id2 f2 l2 t2 ...
                ...
                idn fn ln tn ...
        '''
        validformarts=["TBXTools","freeling","conll"]
        #TODO: Raise exception if not a valid format.
        cf=codecs.open(corpusfile,"r",encoding=encoding)
        if format.lower()=="tbxtools":            
            data=[]
            continserts=0
            for line in cf:
                continserts+=1                
                record=[]
                line=line.rstrip()
                record.append(line)
                data.append(record)
                if continserts==self.maxinserts:
                    self.cur.executemany("INSERT INTO sl_tagged_corpus (tagged_segment) VALUES (?)",data)
                    data=[]
                    continserts=0
                    
            with self.conn:
                self.cur.executemany("INSERT INTO sl_tagged_corpus (tagged_segment) VALUES (?)",data)    
            self.conn.commit()
        elif format.lower()=="freeling":
            data=[]
            continserts=0
            segment=[]
            for line in cf:
                line=line.rstrip()
                if line=="":
                    continserts+=1
                    record=[]
                    record.append(" ".join(segment))
                    data.append(record)
                    if continserts==self.maxinserts:
                        self.cur.executemany("INSERT INTO sl_tagged_corpus (tagged_segment) VALUES (?)",data)
                        data=[]
                        continserts=0
                        data=[]
                        self.conn.commit()
                    segment=[]
                        
                else:
                    camps=line.split()
                    token=camps[0]+"|"+camps[1]+"|"+camps[2]

                    segment.append(token)
            with self.conn:
                self.cur.executemany("INSERT INTO sl_tagged_corpus (tagged_segment) VALUES (?)",data)    
                self.conn.commit()                                                
        elif format.lower()=="conll":
            data=[]
            continserts=0
            segment=[]
            for line in cf:
                line=line.rstrip()
                if line=="":
                    continserts+=1
                    record=[]
                    record.append(" ".join(segment))
                    data.append(record)
                    if continserts==self.maxinserts:
                        self.cur.executemany("INSERT INTO sl_tagged_corpus (tagged_segment) VALUES (?)",self.data)
                        data=[]
                        continserts=0
                        data=[]
                        self.conn.commit()
                    segment=[]
                        
                else:
                    camps=line.split()
                    token=camps[1]+"|"+camps[2]+"|"+camps[3]
                    segment.append(token)                                                
            with self.conn:
                self.cur.executemany("INSERT INTO sl_tagged_corpus (tagged_segment) VALUES (?)",data)    
            self.conn.commit()

    def load_tl_tagged_corpus(self,corpusfile,format="TBXTools",encoding="utf-8"):
        '''Loads the target language tagged corpus. 3 formats are allowed:
        - TBXTools: The internal format used by TBXTools. One tagged segment per line.
                f1|l1|t1|p1 f2|l2|t2|p2 ... fn|ln|tn|pn 
        - Freeling: One token per line and segments separated by blank lines
                f1 l1 t1 p1
                f2 l2 t2 p2
                ...
                fn ln tn pn
        - Conll: One of the output formats guiven by the Standford Core NLP analyzer.  On token per line and segments separated by blank lines
                id1 f1 l1 t1 ...
                id2 f2 l2 t2 ...
                ...
                idn fn ln tn ...
        '''
        validformarts=["TBXTools","freeling","conll"]
        #TODO: Raise exception if not a valid format.
        cf=codecs.open(corpusfile,"r",encoding=encoding)
        if format.lower()=="tbxtools":            
            data=[]
            continserts=0
            for line in cf:
                continserts+=1                
                record=[]
                line=line.rstrip()
                record.append(line)
                data.append(record)
                if continserts==self.maxinserts:
                    self.cur.executemany("INSERT INTO tl_tagged_corpus (tagged_segment) VALUES (?)",data)
                    data=[]
                    continserts=0
                    
            with self.conn:
                self.cur.executemany("INSERT INTO tl_tagged_corpus (tagged_segment) VALUES (?)",data)    
            self.conn.commit()
        elif format.lower()=="freeling":
            data=[]
            continserts=0
            segment=[]
            for line in cf:
                line=line.rstrip()
                if line=="":
                    continserts+=1
                    record=[]
                    record.append(" ".join(segment))
                    data.append(record)
                    if continserts==self.maxinserts:
                        self.cur.executemany("INSERT INTO tl_tagged_corpus (tagged_segment) VALUES (?)",data)
                        data=[]
                        continserts=0
                        data=[]
                        self.conn.commit()
                    segment=[]
                        
                else:
                    camps=line.split()
                    token=camps[0]+"|"+camps[1]+"|"+camps[2]

                    segment.append(token)
            with self.conn:
                self.cur.executemany("INSERT INTO tl_tagged_corpus (tagged_segment) VALUES (?)",data)    
                self.conn.commit()                                                
        elif format.lower()=="conll":
            data=[]
            continserts=0
            segment=[]
            for line in cf:
                line=line.rstrip()
                if line=="":
                    continserts+=1
                    record=[]
                    record.append(" ".join(segment))
                    data.append(record)
                    if continserts==self.maxinserts:
                        self.cur.executemany("INSERT INTO tl_tagged_corpus (tagged_segment) VALUES (?)",self.data)
                        data=[]
                        continserts=0
                        data=[]
                        self.conn.commit()
                    segment=[]
                        
                else:
                    camps=line.split()
                    token=camps[1]+"|"+camps[2]+"|"+camps[3]
                    segment.append(token)                                                
            with self.conn:
                self.cur.executemany("INSERT INTO tl_tagged_corpus (tagged_segment) VALUES (?)",data)    
            self.conn.commit()  
    
    def load_sl_tagged_corpus_c(self,corpusfile,format="TBXTools",encoding="utf-8"):
        '''Loads the source language tagged corpus. 3 formats are allowed:
        - TBXTools: The internal format used by TBXTools. One tagged segment per line.
                f1|l1|t1|p1 f2|l2|t2|p2 ... fn|ln|tn|pn 
        - Freeling: One token per line and segments separated by blank lines
                f1 l1 t1 p1
                f2 l2 t2 p2
                ...
                fn ln tn pn
        - Conll: One of the output formats guiven by the Standford Core NLP analyzer.  On token per line and segments separated by blank lines
                id1 f1 l1 t1 ...
                id2 f2 l2 t2 ...
                ...
                idn fn ln tn ...
        '''
        validformarts=["TBXTools","freeling","conll"]
        #TODO: Raise exception if not a valid format.
        cf=codecs.open(corpusfile,"r",encoding=encoding)
        if format.lower()=="tbxtools":            
            data=[]
            continserts=0
            for line in cf:
                continserts+=1                
                record=[]
                line=line.rstrip()
                record.append(line)
                data.append(record)
                if continserts==self.maxinserts:
                    self.cur.executemany("INSERT INTO sl_tagged_corpus_c (tagged_segment) VALUES (?)",data)
                    data=[]
                    continserts=0
                    
            with self.conn:
                self.cur.executemany("INSERT INTO sl_tagged_corpus_c (tagged_segment) VALUES (?)",data)    
            self.conn.commit()
        elif format.lower()=="freeling":
            data=[]
            continserts=0
            segment=[]
            for line in cf:
                line=line.rstrip()
                if line=="":
                    continserts+=1
                    record=[]
                    record.append(" ".join(segment))
                    data.append(record)
                    if continserts==self.maxinserts:
                        self.cur.executemany("INSERT INTO sl_tagged_corpus_c (tagged_segment) VALUES (?)",data)
                        data=[]
                        continserts=0
                        data=[]
                        self.conn.commit()
                    segment=[]
                        
                else:
                    camps=line.split()
                    token=camps[0]+"|"+camps[1]+"|"+camps[2]

                    segment.append(token)
            with self.conn:
                self.cur.executemany("INSERT INTO sl_tagged_corpus_c (tagged_segment) VALUES (?)",data)    
                self.conn.commit()                                                
        elif format.lower()=="conll":
            data=[]
            continserts=0
            segment=[]
            for line in cf:
                line=line.rstrip()
                if line=="":
                    continserts+=1
                    record=[]
                    record.append(" ".join(segment))
                    data.append(record)
                    if continserts==self.maxinserts:
                        self.cur.executemany("INSERT INTO sl_tagged_corpus_c (tagged_segment) VALUES (?)",self.data)
                        data=[]
                        continserts=0
                        data=[]
                        self.conn.commit()
                    segment=[]
                        
                else:
                    camps=line.split()
                    token=camps[1]+"|"+camps[2]+"|"+camps[3]
                    segment.append(token)                                                
            with self.conn:
                self.cur.executemany("INSERT INTO sl_tagged_corpus_c (tagged_segment) VALUES (?)",data)    
            self.conn.commit()

    def load_tl_tagged_corpus_c(self,corpusfile,format="TBXTools",encoding="utf-8"):
        '''Loads the target language tagged corpus. 3 formats are allowed:
        - TBXTools: The internal format used by TBXTools. One tagged segment per line.
                f1|l1|t1|p1 f2|l2|t2|p2 ... fn|ln|tn|pn 
        - Freeling: One token per line and segments separated by blank lines
                f1 l1 t1 p1
                f2 l2 t2 p2
                ...
                fn ln tn pn
        - Conll: One of the output formats guiven by the Standford Core NLP analyzer.  On token per line and segments separated by blank lines
                id1 f1 l1 t1 ...
                id2 f2 l2 t2 ...
                ...
                idn fn ln tn ...
        '''
        validformarts=["TBXTools","freeling","conll"]
        #TODO: Raise exception if not a valid format.
        cf=codecs.open(corpusfile,"r",encoding=encoding)
        if format.lower()=="tbxtools":            
            data=[]
            continserts=0
            for line in cf:
                continserts+=1                
                record=[]
                line=line.rstrip()
                record.append(line)
                data.append(record)
                if continserts==self.maxinserts:
                    self.cur.executemany("INSERT INTO tl_tagged_corpus_c (tagged_segment) VALUES (?)",data)
                    data=[]
                    continserts=0
                    
            with self.conn:
                self.cur.executemany("INSERT INTO tl_tagged_corpus_c (tagged_segment) VALUES (?)",data)    
            self.conn.commit()
        elif format.lower()=="freeling":
            data=[]
            continserts=0
            segment=[]
            for line in cf:
                line=line.rstrip()
                if line=="":
                    continserts+=1
                    record=[]
                    record.append(" ".join(segment))
                    data.append(record)
                    if continserts==self.maxinserts:
                        self.cur.executemany("INSERT INTO tl_tagged_corpus_c (tagged_segment) VALUES (?)",data)
                        data=[]
                        continserts=0
                        data=[]
                        self.conn.commit()
                    segment=[]
                        
                else:
                    camps=line.split()
                    token=camps[0]+"|"+camps[1]+"|"+camps[2]

                    segment.append(token)
            with self.conn:
                self.cur.executemany("INSERT INTO tl_tagged_corpus_c (tagged_segment) VALUES (?)",data)    
                self.conn.commit()                                                
        elif format.lower()=="conll":
            data=[]
            continserts=0
            segment=[]
            for line in cf:
                line=line.rstrip()
                if line=="":
                    continserts+=1
                    record=[]
                    record.append(" ".join(segment))
                    data.append(record)
                    if continserts==self.maxinserts:
                        self.cur.executemany("INSERT INTO tl_tagged_corpus_c (tagged_segment) VALUES (?)",self.data)
                        data=[]
                        continserts=0
                        data=[]
                        self.conn.commit()
                    segment=[]
                        
                else:
                    camps=line.split()
                    token=camps[1]+"|"+camps[2]+"|"+camps[3]
                    segment.append(token)                                                
            with self.conn:
                self.cur.executemany("INSERT INTO tl_tagged_corpus_c (tagged_segment) VALUES (?)",data)    
            self.conn.commit() 
    
    
    def load_sl_stopwords(self,fitxer,encoding="utf-8"):
        '''Loads the stopwords for the source language.'''
        fc=codecs.open(fitxer,"r",encoding)
        data=[]
        record=[]
        while 1:
            linia=fc.readline()
            if not linia:
                break 
            linia=linia.rstrip()
            record.append(linia)
            data.append(record)
            record=[]
        
        for punct in self.punctuation:
            record.append(punct)
            data.append(record)
            record=[]
        with self.conn:
            self.cur.executemany("INSERT INTO sl_stopwords (sl_stopword) VALUES (?)",data)  
            
    def load_tl_stopwords(self,fitxer,encoding="utf-8"):
        '''Loads the stopwords for the target language.'''
        fc=codecs.open(fitxer,"r",encoding)
        data=[]
        record=[]
        while 1:
            linia=fc.readline()
            if not linia:
                break 
            linia=linia.rstrip()
            record.append(linia)
            data.append(record)
            record=[]
        
        for punct in self.punctuation:
            record.append(punct)
            data.append(record)
            record=[]
        with self.conn:
            self.cur.executemany("INSERT INTO tl_stopwords (tl_stopword) VALUES (?)",data) 

    def load_sl_inner_stopwords(self,fitxer,encoding="utf-8"):
        '''Loads the stopwords for the source language.'''
        fc=codecs.open(fitxer,"r",encoding)
        data=[]
        record=[]
        while 1:
            linia=fc.readline()
            if not linia:
                break 
            linia=linia.rstrip()
            record.append(linia)
            data.append(record)
            record=[]        
        for punct in self.punctuation:
            record.append(punct)
            data.append(record)
            record=[]
        with self.conn:
            self.cur.executemany("INSERT INTO sl_inner_stopwords (sl_inner_stopword) VALUES (?)",data)  
            
    def load_tl_inner_stopwords(self,fitxer,encoding="utf-8"):
        '''Loads the inner stopwords for the target language.'''
        fc=codecs.open(fitxer,"r",encoding)
        data=[]
        record=[]
        while 1:
            linia=fc.readline()
            if not linia:
                break 
            linia=linia.rstrip()
            record.append(linia)
            data.append(record)
            record=[]        
        for punct in self.punctuation:
            record.append(punct)
            data.append(record)
            record=[]
        with self.conn:
            self.cur.executemany("INSERT INTO tl_inner_stopwords (tl_inner_stopword) VALUES (?)",data)  
    
    #evaluation terms
    def load_evaluation_terms_tabtxt(self,arxiu,encoding="utf-8",nmin=0,nmax=1000):
        '''Loads the evaluation terms from a tabulated text.'''
        cf=codecs.open(arxiu,"r",encoding=encoding)
        data=[]
        continserts=0
        for line in cf:
            line=line.rstrip()
            continserts+=1            
            record=[]
            line=line.rstrip()
            camps=line.split("\t")
            if self.specificSLtokenizer:
                tokens=self.SLtokenizer.tokenize(camps[0]).split()
            else:
                tokens=camps[0].split()
            if len(camps)==1:
                if len(tokens)>=nmin and len(tokens)<=nmax:
                    record.append(camps[0])
                    record.append("_")
                    data.append(record)
            elif len(camps)>1:
                if len(tokens)>=nmin and len(tokens)<=nmax:
                    record.append(camps[0])
                    record.append(camps[1])
                    data.append(record)
            if continserts==self.maxinserts: 
                self.cur.executemany("INSERT INTO evaluation_terms (sl_term,tl_term) VALUES (?,?)",data)
                data=[]
                continserts=0
        with self.conn:
            self.cur.executemany("INSERT INTO evaluation_terms (sl_term,tl_term) VALUES (?,?)",data)
        self.conn.commit()
        
    def load_evaluation_terms_tbx(self,arxiu,sl_code="",tl_code="",encoding="utf-8",nmin=0,nmax=1000):
        '''Loads the evaluation terms from a TBX file.'''
        slcodes=[]
        for slc in sl_code.split(","):
            slcodes.append(slc.strip())
        tlcodes=[]
        for tlc in tl_code.split(","):
            tlcodes.append(tlc.strip())
        data=[]
        slterm=[]
        tlterm=[]
        lang=""
        for event, elem in etree.iterparse(arxiu,events=("start", "end")):
            tag=elem.tag.replace(self.namespace(elem),"")
            if event=="end" and tag in ["conceptEntry","termEntry"]:
                if len(slterm)>0 and len(tlterm)>0:
                    record=[]
                    for slt in slterm:
                        if self.specificSLtokenizer:
                            tokens=self.SLtokenizer.tokenize(slt).split()
                        else:
                            tokens=slt.split()
                        if len(tokens)>=nmin and len(tokens)<=nmax:
                            tlt=", ".join(tlterm)
                            record.append(slt)
                            record.append(tlt)
                            data.append(record)  
                            record=[]
                    slterm=[]
                    tlterm=[]
            elif event=="start" and tag=="langSec":
                if elem.attrib["{http://www.w3.org/XML/1998/namespace}lang"] in slcodes:
                    lang=elem.attrib["{http://www.w3.org/XML/1998/namespace}lang"]
                if elem.attrib["{http://www.w3.org/XML/1998/namespace}lang"] in tlcodes:
                    lang=elem.attrib["{http://www.w3.org/XML/1998/namespace}lang"]
            elif event=="start" and tag=="term":
                if lang in slcodes: slterm.append("".join(elem.itertext()).lstrip().rstrip())
                elif lang in tlcodes: tlterm.append("".join(elem.itertext()).lstrip().rstrip())
        self.cur.executemany("INSERT INTO evaluation_terms (sl_term,tl_term) VALUES (?,?)",data)   
        self.conn.commit()
    #
    def load_validated_terms(self,terms):
        """Load a list of tuples containig source-target terms)."""
        data=[]
        for tupleTerms in terms:
            record=[]
            slterm=tupleTerms[0]
            tlterm=tupleTerms[1]
            record.append(slterm)
            record.append(tlterm)
            data.append(record)
        self.cur.executemany("INSERT INTO validated_terms (sl_term,tl_term) VALUES (?,?)",data)   
        self.conn.commit()
    def get_validated_terms(self):
        self.cur.execute("SELECT sl_term, tl_term FROM validated_terms;")
        validatedterms=[]
        source_terms=[]
        target_terms=[]
        for s in self.cur.fetchall():
            record=[]
            record.append(s[0])
            record.append(s[1])
            validatedterms.append(record)
        return(validatedterms)
        
    #reference_terms
    def load_reference_terms_tabtxt(self,arxiu,encoding="utf-8",nmin=0,nmax=1000, reverse=False):
        '''Loads the reference terms from a tabulated text.'''
        cf=codecs.open(arxiu,"r",encoding=encoding)
        data=[]
        continserts=0
        for line in cf:
            line=line.rstrip()
            continserts+=1            
            record=[]
            line=line.rstrip()
            camps=line.split("\t")
            if self.specificSLtokenizer:
                tokens=self.SLtokenizer.tokenize(camps[0]).split()
            else:
                tokens=camps[0].split()
            if len(camps)==1:
                if len(tokens)>=nmin and len(tokens)<=nmax:
                    record.append(camps[0])
                    record.append("_")
                    data.append(record)
            elif len(camps)>1:
                if len(tokens)>=nmin and len(tokens)<=nmax:
                    record.append(camps[0])
                    record.append(camps[1])
                    data.append(record)
            if continserts==self.maxinserts: 
                if reverse:
                    self.cur.executemany("INSERT INTO reference_terms (tl_term,sl_term) VALUES (?,?)",data)
                else:
                    self.cur.executemany("INSERT INTO reference_terms (sl_term,tl_term) VALUES (?,?)",data)
                data=[]
                continserts=0
        with self.conn:
            if reverse:
                self.cur.executemany("INSERT INTO reference_terms (tl_term,sl_term) VALUES (?,?)",data)
            else:
                self.cur.executemany("INSERT INTO reference_terms (sl_term,tl_term) VALUES (?,?)",data)
        self.conn.commit()
        
    def load_reference_terms_tbx(self,arxiu,sl_code="",tl_code="",encoding="utf-8",nmin=0,nmax=1000):
        '''Loads the evaluation terms from a TBX file.'''
        slcodes=[]
        for slc in sl_code.split(","):
            slcodes.append(slc.strip())
        tlcodes=[]
        for tlc in tl_code.split(","):
            tlcodes.append(tlc.strip())
        data=[]
        slterm=[]
        tlterm=[]
        lang=""
        for event, elem in etree.iterparse(arxiu,events=("start", "end")):
            tag=elem.tag.replace(self.namespace(elem),"")
            if event=="end" and tag in ["conceptEntry","termEntry"]:
                if len(slterm)>0 and len(tlterm)>0:
                    record=[]
                    for slt in slterm:
                        if self.specificSLtokenizer:
                            tokens=self.SLtokenizer.tokenize(slt).split()
                        else:
                            tokens=slt.split()
                        if len(tokens)>=nmin and len(tokens)<=nmax:
                            tlt=", ".join(tlterm)
                            record.append(slt)
                            record.append(tlt)
                            data.append(record)  
                            record=[]
                    slterm=[]
                    tlterm=[]
            elif event=="start" and tag=="langSec":
                if elem.attrib["{http://www.w3.org/XML/1998/namespace}lang"] in slcodes:
                    lang=elem.attrib["{http://www.w3.org/XML/1998/namespace}lang"]
                if elem.attrib["{http://www.w3.org/XML/1998/namespace}lang"] in tlcodes:
                    lang=elem.attrib["{http://www.w3.org/XML/1998/namespace}lang"]
            elif event=="start" and tag=="term":
                if lang in slcodes: slterm.append("".join(elem.itertext()).lstrip().rstrip())
                elif lang in tlcodes: tlterm.append("".join(elem.itertext()).lstrip().rstrip())
        self.cur.executemany("INSERT INTO reference_terms (sl_term,tl_term) VALUES (?,?)",data)   
        self.conn.commit()
    
    def load_reference_terms_csv(self,arxiu,encoding="utf-8",nmin=0,nmax=1000,CSVdelimiter=",",CSVquotechar=None,CSVescapechar=None,CSVSLTerm=1,CSVTLTerm=2):
        csv_file=codecs.open(arxiu,"r",encoding=encoding)
        csv_reader = csv.reader(csv_file, delimiter=",", quotechar=CSVquotechar, escapechar=CSVescapechar)
        record=[]
        data=[]
        for row in csv_reader:
            record.append(row[CSVSLTerm-1])
            record.append(row[CSVTLTerm-1])
            data.append(record)  
            record=[]
        self.cur.executemany("INSERT INTO reference_terms (sl_term,tl_term) VALUES (?,?)",data)   
        self.conn.commit()
    def load_reference_terms_excel(self,file,nmin=0,nmax=1000,sheet_name=1,first_row=1,sourceColumn="A",targetColumn="B"):
        workbook = load_workbook(filename=file)
        data=[]
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            for row in sheet.rows:
                source=""
                target=""
                record=[]
                for cell in row:
                    
                    if isinstance(cell, openpyxl.cell.cell.MergedCell):
                        # Skip this cell
                        continue
                    if cell.column_letter==sourceColumn: 
                        source=cell.value
                    elif cell.column_letter==targetColumn: 
                        target=cell.value
                if not source=="" and not target=="":
                    record.append(source)
                    record.append(target)  
                    data.append(record)
        self.cur.executemany("INSERT INTO reference_terms (sl_term,tl_term) VALUES (?,?)",data)
        self.conn.commit()

    #compoundify_terms_sl
    def load_compoundify_terms_sl_txt(self,arxiu,encoding="utf-8",nmin=0,nmax=1000):
        '''Loads the compoundify terms for the source language from a text file (one term per line).'''
        cf=codecs.open(arxiu,"r",encoding=encoding)
        data=[]
        continserts=0
        for line in cf:
            line=line.rstrip()
            continserts+=1            
            record=[]
            line=line.rstrip()
            camps=line.split("\t")
            if self.specificSLtokenizer:
                tokens=self.SLtokenizer.tokenize(camps[0]).split()
            else:
                tokens=camps[0].split()
            if len(tokens)>=nmin and len(tokens)<=nmax:
                record.append(camps[0])
                data.append(record)
            if continserts==self.maxinserts: 
                self.cur.executemany("INSERT INTO compoundify_terms_sl (term) VALUES (?)",data)
                data=[]
                continserts=0
        with self.conn:
            self.cur.executemany("INSERT INTO compoundify_terms_sl (term) VALUES (?)",data)
        self.conn.commit()
        
    def load_compoundify_terms_sl_tbx(self,arxiu,code="",encoding="utf-8",nmin=0,nmax=1000):
        '''Loads the compoundify terms for the source language from a TBX file.'''
        codes=[]
        for slc in code.split(","):
            codes.append(slc.strip())
        data=[]
        term=[]
        lang=""
        for event, elem in etree.iterparse(arxiu,events=("start", "end")):
            tag=elem.tag.replace(self.namespace(elem),"")
            if event=="end" and tag in ["conceptEntry","termEntry"]:
                if len(term)>0 and lang in codes:
                    record=[]
                    for slt in term:
                        if self.specificSLtokenizer:
                            tokens=self.SLtokenizer.tokenize(slt).split()
                        else:
                            tokens=slt.split()
                        if len(tokens)>=nmin and len(tokens)<=nmax:
                            record.append(slt)
                            data.append(record)  
                            record=[]
                    term=[]
            elif event=="start" and tag=="langSec":
                if elem.attrib["{http://www.w3.org/XML/1998/namespace}lang"] in codes:
                    lang=elem.attrib["{http://www.w3.org/XML/1998/namespace}lang"]
                else:
                    lang=""
            elif event=="start" and tag=="term":
                if lang in codes: 
                    term.append("".join(elem.itertext()).lstrip().rstrip())
        self.cur.executemany("INSERT INTO compoundify_terms_sl (term) VALUES (?)",data)   
        self.conn.commit()
        
    #compoundify_terms_tl
    def load_compoundify_terms_tl_txt(self,arxiu,encoding="utf-8",nmin=0,nmax=1000):
        '''Loads the compoundify terms for the target language from a text file (one term per line).'''
        cf=codecs.open(arxiu,"r",encoding=encoding)
        data=[]
        continserts=0
        for line in cf:
            line=line.rstrip()
            continserts+=1            
            record=[]
            line=line.rstrip()
            camps=line.split("\t")
            if self.specificSLtokenizer:
                tokens=self.SLtokenizer.tokenize(camps[0]).split()
            else:
                tokens=camps[0].split()
            if len(tokens)>=nmin and len(tokens)<=nmax:
                record.append(camps[0])
                data.append(record)
            if continserts==self.maxinserts: 
                self.cur.executemany("INSERT INTO compoundify_terms_tl (term) VALUES (?)",data)
                data=[]
                continserts=0
        with self.conn:
            self.cur.executemany("INSERT INTO compoundify_terms_tl (term) VALUES (?)",data)
        self.conn.commit()
        
    def load_compoundify_terms_tl_tbx(self,arxiu,code="",encoding="utf-8",nmin=0,nmax=1000):
        '''Loads the compoundify terms for the target language from a TBX file.'''
        codes=[]
        for slc in code.split(","):
            codes.append(slc.strip())
        data=[]
        term=[]
        lang=""
        for event, elem in etree.iterparse(arxiu,events=("start", "end")):
            tag=elem.tag.replace(self.namespace(elem),"")
            if event=="end" and tag in ["conceptEntry","termEntry"]:
                if len(term)>0 and lang in codes:
                    record=[]
                    for slt in term:
                        if self.specificSLtokenizer:
                            tokens=self.SLtokenizer.tokenize(slt).split()
                        else:
                            tokens=slt.split()
                        if len(tokens)>=nmin and len(tokens)<=nmax:
                            record.append(slt)
                            data.append(record)  
                            record=[]
                    term=[]
            elif event=="start" and tag=="langSec":
                if elem.attrib["{http://www.w3.org/XML/1998/namespace}lang"] in codes:
                    lang=elem.attrib["{http://www.w3.org/XML/1998/namespace}lang"]
                else:
                    lang=""
            elif event=="start" and tag=="term":
                if lang in codes: 
                    term.append("".join(elem.itertext()).lstrip().rstrip())
        self.cur.executemany("INSERT INTO compoundify_terms_tl (term) VALUES (?)",data)   
        self.conn.commit()
        
    #tsr terms
    
    def load_tsr_terms_txt(self,arxiu,encoding="utf-8",nmin=0,nmax=1000):
        '''Loads the TSR terms from a text file (one term per line).'''
        cf=codecs.open(arxiu,"r",encoding=encoding)
        data=[]
        continserts=0
        for line in cf:
            line=line.rstrip()
            continserts+=1            
            record=[]
            line=line.rstrip()
            camps=line.split("\t")
            if self.specificSLtokenizer:
                tokens=self.SLtokenizer.tokenize(camps[0]).split()
            else:
                tokens=camps[0].split()
            if len(tokens)>=nmin and len(tokens)<=nmax:
                record.append(camps[0])
                data.append(record)
            if continserts==self.maxinserts: 
                self.cur.executemany("INSERT INTO tsr_terms (term) VALUES (?)",data)
                data=[]
                continserts=0
        with self.conn:
            self.cur.executemany("INSERT INTO tsr_terms (term) VALUES (?)",data)
        self.conn.commit()
        
    def load_tosearch_terms(self,SLterms,encoding="utf-8",nmin=0,nmax=1000):
        '''Loads the TSR terms from a string, text file (one term per line) or list.'''
        tofind=[]
        if isinstance(SLterms, str):
            if os.path.exists(SLterms):
                entrada=codecs.open(SLterms)
                for linia in entrada:
                    linia=linia.rstrip()
                    tofind.append(linia)
                entrada.close()
            else:
                tofind.append(SLterms)
        elif isinstance(SLterms, list):
            tofind.extend(SLterms)
        data=[]
        continserts=0
        for term in tofind:
            continserts+=1            
            record=[]
            if self.specificSLtokenizer:
                tokens=self.SLtokenizer.tokenize(term)
            else:
                tokens=term.split()
            if len(tokens)>=nmin and len(tokens)<=nmax:
                record.append(term)
                data.append(record)
            if continserts==self.maxinserts: 
                self.cur.executemany("INSERT INTO tosearch_terms (term) VALUES (?)",data)
                data=[]
                continserts=0
        with self.conn:
            self.cur.executemany("INSERT INTO tosearch_terms (term) VALUES (?)",data)
        self.conn.commit()
        
    def load_tsr_terms_tbx(self,arxiu,code="",encoding="utf-8",nmin=0,nmax=1000):
        '''Loads the TSR terms from a TBX file.'''
        codes=[]
        for slc in code.split(","):
            codes.append(slc.strip())
        data=[]
        term=[]
        lang=""
        for event, elem in etree.iterparse(arxiu,events=("start", "end")):
            tag=elem.tag.replace(self.namespace(elem),"")
            if event=="end" and tag in ["conceptEntry","termEntry"]:
                if len(term)>0 and lang in codes:
                    record=[]
                    for slt in term:
                        if self.specificSLtokenizer:
                            tokens=self.SLtokenizer.tokenize(slt).split()
                        else:
                            tokens=slt.split()
                        if len(tokens)>=nmin and len(tokens)<=nmax:
                            record.append(slt)
                            data.append(record)  
                            record=[]
                    term=[]
            elif event=="start" and tag=="langSec":
                if elem.attrib["{http://www.w3.org/XML/1998/namespace}lang"] in codes:
                    lang=elem.attrib["{http://www.w3.org/XML/1998/namespace}lang"]
                else:
                    lang=""
            elif event=="start" and tag=="term":
                if lang in codes: 
                    term.append("".join(elem.itertext()).lstrip().rstrip())
        self.cur.executemany("INSERT INTO tsr_terms (term) VALUES (?)",data)   
        self.conn.commit()
        
    #exclusion_terms
    
    def load_exclusion_terms_tabtxt(self,arxiu,encoding="utf-8",nmin=0,nmax=1000):
        '''Loads the exclusion terms from a tabulated text.'''
        cf=codecs.open(arxiu,"r",encoding=encoding)
        data=[]
        continserts=0
        for line in cf:
            line=line.rstrip()
            continserts+=1            
            record=[]
            line=line.rstrip()
            camps=line.split("\t")
            if self.specificSLtokenizer:
                tokens=self.SLtokenizer.tokenize(camps[0]).split()
            else:
                tokens=camps[0].split()
            if len(camps)==1:
                if len(tokens)>=nmin and len(tokens)<=nmax:
                    record.append(camps[0])
                    record.append("_")
                    data.append(record)
            elif len(camps)>1:
                if len(tokens)>=nmin and len(tokens)<=nmax:
                    record.append(camps[0])
                    record.append(camps[1])
                    data.append(record)
            if continserts==self.maxinserts: 
                self.cur.executemany("INSERT INTO exclusion_terms (sl_term,tl_term) VALUES (?,?)",data)
                data=[]
                continserts=0
        with self.conn:
            self.cur.executemany("INSERT INTO exclusion_terms (sl_term,tl_term) VALUES (?,?)",data)
        self.conn.commit()
        
    def load_exclusion_terms_tbx(self,arxiu,sl_code="",tl_code="",encoding="utf-8",nmin=0,nmax=1000):
        '''Loads the exclusion terms from a TBX file.'''
        slcodes=[]
        for slc in sl_code.split(","):
            slcodes.append(slc.strip())
        tlcodes=[]
        for tlc in tl_code.split(","):
            tlcodes.append(tlc.strip())
        data=[]
        slterm=[]
        tlterm=[]
        lang=""
        for event, elem in etree.iterparse(arxiu,events=("start", "end")):
            tag=elem.tag.replace(self.namespace(elem),"")
            if event=="end" and tag in ["conceptEntry","termEntry"]:
                if len(slterm)>0 and len(tlterm)>0:
                    record=[]
                    for slt in slterm:
                        if self.specificSLtokenizer:
                            tokens=self.SLtokenizer.tokenize(slt).split()
                        else:
                            tokens=slt.split()
                        if len(tokens)>=nmin and len(tokens)<=nmax:
                            tlt=", ".join(tlterm)
                            record.append(slt)
                            record.append(tlt)
                            data.append(record)  
                            record=[]
                    slterm=[]
                    tlterm=[]
            elif event=="start" and tag=="langSec":
                if elem.attrib["{http://www.w3.org/XML/1998/namespace}lang"] in slcodes:
                    lang=elem.attrib["{http://www.w3.org/XML/1998/namespace}lang"]
                if elem.attrib["{http://www.w3.org/XML/1998/namespace}lang"] in tlcodes:
                    lang=elem.attrib["{http://www.w3.org/XML/1998/namespace}lang"]
            elif event=="start" and tag=="term":
                if lang in slcodes: slterm.append("".join(elem.itertext()).lstrip().rstrip())
                elif lang in tlcodes: tlterm.append("".join(elem.itertext()).lstrip().rstrip())
        self.cur.executemany("INSERT INTO exclusion_terms (sl_term,tl_term) VALUES (?,?)",data)   
        self.conn.commit()
        
    #EXCLUSION NO TERMS
    def load_exclusion_noterms_tabtxt(self,arxiu,encoding="utf-8",nmin=0,nmax=1000):
        '''Loads the exclusion no terms from a tabulated text.'''
        cf=codecs.open(arxiu,"r",encoding=encoding)
        data=[]
        continserts=0
        for line in cf:
            line=line.rstrip()
            continserts+=1            
            record=[]
            line=line.rstrip()
            camps=line.split("\t")
            if self.specificSLtokenizer:
                tokens=self.SLtokenizer.tokenize(camps[0]).split()
            else:
                tokens=camps[0].split()
            if len(camps)==1:
                if len(tokens)>=nmin and len(tokens)<=nmax:
                    record.append(camps[0])
                    record.append("_")
                    data.append(record)
            elif len(camps)>1:
                if len(tokens)>=nmin and len(tokens)<=nmax:
                    record.append(camps[0])
                    record.append(camps[1])
                    data.append(record)
            if continserts==self.maxinserts: 
                self.cur.executemany("INSERT INTO exclusion_noterms (sl_term,tl_term) VALUES (?,?)",data)
                data=[]
                continserts=0
        with self.conn:
            self.cur.executemany("INSERT INTO exclusion_noterms (sl_term,tl_term) VALUES (?,?)",data)
    
    
    
    def namespace(self,element):
        m = re.match(r'\{.*\}', element.tag)
        return m.group(0) if m else ''
        
    def find_translation_reference_terms(self,term):
        self.cur.execute("SELECT tl_term FROM reference_terms where sl_term='"+str(term)+"'")
        tlterms=[]
        for self.s in self.cur.fetchall():
            tlterms.append(self.s[0])
        if len(tlterms)>0:
            return(", ".join(tlterms))
        else:
            return(None)
        
    
    def load_sl_exclusion_regexps(self,arxiu,encoding="utf-8"):
        '''Loads the exclusion regular expressions for the source language.'''
        cf=codecs.open(arxiu,"r",encoding=encoding)
        data=[]
        for line in cf:
            line=line.rstrip()
            record=[]
            record.append(line)
            data.append(record)
            
        with self.conn:
            self.cur.executemany('INSERT INTO sl_exclusion_regexps (sl_exclusion_regexp) VALUES (?)',data)
            
    def load_tl_exclusion_regexps(self,arxiu,encoding="utf-8"):
        '''Loads the exclusion regular expressions for the target language.'''
        cf=codecs.open(arxiu,"r",encoding=encoding)
        data=[]
        for line in cf:
            line=line.rstrip()
            record=[]
            record.append(line)
            data.append(record)
            
        with self.conn:
            self.cur.executemany('INSERT INTO tl_exclusion_regexps (sl_exclusion_regexp) VALUES (?)',data)

   
    def show_term_candidates(self,limit=-1,minfreq=2, minmeasure=-1, show_frequency=True, show_measure=False, mark_eval=False, verbose=False):
        '''Shows the term candidates in the screen.'''
        measure=0
        knownterms=[]
        knownoterms=[]
        with self.conn:
            self.cur.execute("SELECT sl_term FROM exclusion_terms")
            for s in self.cur.fetchall():
                knownterms.append(s[0])
        with self.conn:
            self.cur.execute("SELECT sl_term FROM exclusion_noterms")
            for s in self.cur.fetchall():
                knownnoterms.append(s[0])
        with self.conn:
            self.cur.execute("SELECT frequency,value,n,candidate FROM term_candidates order by value desc, frequency desc, random() limit "+str(limit))
            for s in self.cur.fetchall():
                frequency=s[0]
                if s[1]==None:
                    measure==0
                else:
                    measure=s[1]
                n=s[2]
                candidate=s[3]
                if n>=n_min and n<=n_max and not candidate in knownterms and not candidate in knownoterms:
                    if mark_eval:
                        if candidate in evaluation_terms:
                            candidate="*"+candidate
                    if show_frequency and not show_measure:
                        cadena=str(frequency)+"\t"+candidate
                    if not show_frequency and show_measure:
                        cadena=str(measure)+"\t"+candidate
                    if show_measure and show_frequency:
                        cadena=str(frequency)+"\t"+str(measure)+"\t"+candidate
                    else:
                        cadena=candidate
                    print(cadena)

    def select_unigrams(self,file,position=-1,verbose=True):
        sunigrams=codecs.open(file,"w",encoding="utf-8")
        unigrams={}
        self.cur.execute("SELECT frequency,candidate FROM term_candidates order by value desc, frequency desc, random()")
        #self.cur.execute("SELECT frequency,value,n,candidate FROM term_candidates order by n desc limit "+str(limit))
        for s in self.cur.fetchall():
            frequency=s[0]
            candidate=s[1].split()[position]
            if candidate in unigrams:
                unigrams[candidate]+=frequency
            else:
                unigrams[candidate]=frequency
        #for self.candidate in self.unigrams:
        #    print(self.unigrams[self.candidate],self.candidate)
        data=[]
        for candidate in sorted(unigrams, key=unigrams.get, reverse=True):
            
            cadena=str(unigrams[candidate])+"\t"+candidate
            #if self.verbose: print(cadena)
            record=[]
            record.append(candidate)
            record.append(1)
            record.append(unigrams[candidate])
            record.append("freq")
            record.append(unigrams[candidate])
            data.append(record)
            sunigrams.write(cadena+"\n")
          
        with self.conn:
            self.cur.executemany("INSERT INTO term_candidates (candidate, n, frequency, measure, value) VALUES (?,?,?,?,?)",data)        
            self.conn.commit()


    def save_term_candidates(self,outfile,limit=-1,minfreq=2, minmeasure=-1, show_frequency=True, show_measure=False, mark_eval=False, verbose=False):
        '''Saves the term candidates in a file.'''
        sortida=codecs.open(outfile,"w",encoding="utf-8")
        measure=0
        knownterms=[]
        knownnoterms=[]
        with self.conn:
            self.cur.execute("SELECT sl_term FROM exclusion_terms")
            for s in self.cur.fetchall():
                knownterms.append(s[0])
        with self.conn:
            self.cur.execute("SELECT sl_term FROM exclusion_noterms")
            for s in self.cur.fetchall():
                knownnoterms.append(s[0])
        with self.conn:
            self.cur.execute("SELECT frequency,value,n,candidate FROM term_candidates order by value desc, frequency desc, random() limit "+str(limit))
            for s in self.cur.fetchall():
                frequency=s[0]
                if s[1]==None:
                    measure==0
                else:
                    measure=s[1]
                n=s[2]
                candidate=s[3]
                if not candidate in knownterms and not candidate in knownnoterms:
                    if mark_eval:
                        if candidate in evaluation_terms:
                            candidate="*"+candidate
                    if show_measure and not show_frequency:
                        cadena=str(measure)+"\t"+candidate
                    elif show_frequency and not show_measure:
                        cadena=str(frequency)+"\t"+candidate
                    elif show_frequency and show_measure:
                        cadena=str(frequency)+"\t"+str(measure)+"\t"+candidate
                    else:
                        cadena=candidate
                    if verbose:
                        print(cadena)
                    sortida.write(cadena+"\n")
                    
    #STATISTICAL TERM EXTRACTION
    
    def ngram_calculation (self,nmin,nmax,minfreq=2,corpus="sl_corpus"):
        '''Performs the calculation of ngrams.'''
        ngramsFD=FreqDist()
        tokensFD=FreqDist()
        n_nmin=nmin
        n_max=nmax
            
        with self.conn:
            if corpus=="sl_corpus":
                self.cur.execute('SELECT segment from sl_corpus')
            elif corpus=="tl_corpus":
                self.cur.execute('SELECT segment from tl_corpus')
            for s in self.cur.fetchall():
                segment=s[0]
                for n in range(nmin,nmax+1): #we DON'T calculate one order bigger in order to detect nested candidates
                    if self.specificSLtokenizer:
                        tokens=self.SLtokenizer.tokenize(segment).split()
                    else:
                        tokens=segment.split()
                    ngs=ngrams(tokens, n)
                    for ng in ngs:
                        ngramsFD[ng]+=1
                for token in tokens:
                    tokensFD[token]+=1
                       
        data=[]                
        for c in ngramsFD.most_common():
            if c[1]>=minfreq:
                record=[]
                record.append(" ".join(c[0]))            
                record.append(len(c[0]))
                record.append(c[1])   
                data.append(record)
        with self.conn:
            self.cur.executemany("INSERT INTO ngrams (ngram, n, frequency) VALUES (?,?,?)",data) 
            self.conn.commit()
            
        data=[]                
        for c in tokensFD.most_common():
            record=[]
            record.append(c[0])            
            record.append(c[1])   
            data.append(record)
        with self.conn:
            self.cur.executemany("INSERT INTO tokens (token, frequency) VALUES (?,?)",data) 
            self.conn.commit()
            
    def statistical_term_extraction(self,minfreq=2,corpus="sl_corpus"):
        '''Performs an statistical term extraction using the extracted ngrams (ngram_calculation should be executed first). Loading stop-words is advisable. '''
        self.cur.execute("DELETE FROM term_candidates")
        self.conn.commit()
        stopwords=[]
        with self.conn:
            if corpus=="sl_corpus":
                self.cur.execute("SELECT sl_stopword FROM sl_stopwords")
            elif corpus=="tl_corpus":
                self.cur.execute("SELECT tl_stopword FROM tl_stopwords")
            for s in self.cur.fetchall():
                stopwords.append(s[0])
                
        inner_stopwords=[]
        with self.conn:
            if corpus=="sl_corpus":
                self.cur.execute("SELECT sl_inner_stopword FROM sl_inner_stopwords")
            elif corpus=="tl_corpus":
                self.cur.execute("SELECT tl_inner_stopword FROM tl_inner_stopwords")
            for s in self.cur.fetchall():
                inner_stopwords.append(s[0])
        
        self.cur.execute("SELECT ngram, n, frequency FROM ngrams order by frequency desc")
        results=self.cur.fetchall()
        data=[] 
        for a in results:
            if corpus=="sl_corpus":
                if self.specificSLtokenizer:
                    ng=self.SLtokenizer.tokenize(a[0]).split()
                else:
                    ng=a[0].split()
            if corpus=="tl_corpus":
                if self.specificTLtokenizer:
                    
                    ng=self.TLtokenizer.tokenize(a[0]).split()
                else:
                    ng=a[0].split()
            include=True
            if ng[0].lower() in stopwords: include=False
            if ng[-1].lower() in stopwords: include=False
            for i in range(1,len(ng)):
                if ng[i].lower() in inner_stopwords:
                    include=False
            if include:
                record=[]
                record.append(a[0])            
                record.append(a[1])
                record.append(a[2])
                record.append("freq")
                record.append(a[2])   
                data.append(record)
            if a[2]<minfreq:
                break
        with self.conn:
            self.cur.executemany("INSERT INTO term_candidates (candidate, n, frequency, measure, value) VALUES (?,?,?,?,?)",data)        
            self.conn.commit()
    
    def loadSLtokenizer(self, tokenizer):
        if not tokenizer.endswith(".py"): tokenizer=tokenizer+".py"
        spec = importlib.util.spec_from_file_location('', tokenizer)
        tokenizermod = importlib.util.module_from_spec(spec)
        spec.loader.exec_module(tokenizermod)
        self.SLtokenizer=tokenizermod.Tokenizer()
        self.specificSLtokenizer=True
        
    def unloadSLtokenizer(self):
        self.SLtokenizer=None
        self.specificSLtokenizer=False
        
    def loadTLtokenizer(self, tokenizer):
        if not tokenizer.endswith(".py"): tokenizer=tokenizer+".py"
        spec = importlib.util.spec_from_file_location('', tokenizer)
        tokenizermod = importlib.util.module_from_spec(spec)
        spec.loader.exec_module(tokenizermod)
        self.TLtokenizer=tokenizermod.Tokenizer()
        self.specificTLtokenizer=True
        
    def unloadSLtokenizer(self):
        self.TLtokenizer=None
        self.specificTLtokenizer=False
            
    def statistical_term_extraction_by_segment(self, segment, minlocalfreq=1, minglobalfreq=2, maxcandidates=2, nmin=1, nmax=4):
        '''Performs an statistical term extraction over a single segment using the extracted ngrams (ngram_calculation should be executed first) Loading stop-words is advisable. '''
        ngramsFD=FreqDist()
        sl_stopword=[]
        with self.conn:
            self.cur.execute("SELECT sl_stopword FROM sl_stopwords")
            for s in self.cur.fetchall():
                sl_stopword.append(s[0])
                
        sl_inner_stopwords=[]
        with self.conn:
            self.cur.execute("SELECT sl_inner_stopword FROM sl_inner_stopwords")
            for s in self.cur.fetchall():
                sl_inner_stopwords.append(s[0])
        
        for n in range(nmin,nmax+1):
            if self.specificSLtokenizer:
                tokens=self.SLtokenizer.tokenize(segment).split()
            else:
                tokens=segment.split()
            ngs=ngrams(tokens, n)
            for ng in ngs:
                include=True
                
                if ng[0].lower() in self.sl_stopwords: include=False
                if ng[-1].lower() in self.sl_stopwords: include=False
                for i in range(1,len(ng)):
                    if ng[i].lower() in self.sl_inner_stopwords:
                        include=False
                if include: ngramsFD[" ".join(ng)]+=1
                
        for ng in ngramsFD.most_common():
            print(ng)

    def case_normalization(self,verbose=False):
        '''
        Performs case normalization. If a capitalized term exists as non-capitalized, the capitalized one will be deleted and the frequency of the non-capitalized one will be increased by the frequency of the capitalized.
        '''
        self.cur.execute("SELECT candidate,frequency FROM term_candidates order by frequency desc")
        results=self.cur.fetchall()
        auxiliar={}
        for r in results:
            auxiliar[r[0]]=r[1]
        for a in results:
            if not a[0]==a[0].lower() and a[0].lower() in auxiliar:
                terma=a[0]
                termb=a[0].lower()
                freqa=a[1]
                freqb=auxiliar[termb]
                n=len(termb.split())
                freqtotal=freqa+freqb
                if verbose:
                    print(terma,freqa,"-->",termb,freqb,"-->",freqtotal)
                self.cur.execute('DELETE FROM term_candidates WHERE candidate=?', (terma,))
                self.cur.execute('DELETE FROM term_candidates WHERE candidate=?', (termb,))
                self.cur.execute("INSERT INTO term_candidates (candidate, n, frequency, measure, value) VALUES (?,?,?,?,?)",(termb,n,freqtotal,"freq",freqtotal))
        self.conn.commit()

    def nest_normalization(self,percent=10,verbose=False):
        '''
        Performs a normalization of nested term candidates. If an n-gram candidate A is contained in a n+1 candidate B and freq(A)==freq(B) or they are close values (determined by the percent parameter, A is deleted B remains as it is)
        '''
        self.cur.execute("SELECT candidate,frequency,n FROM term_candidates order by frequency desc")
        results=self.cur.fetchall()
        for a in results:
            ta=a[0]
            fa=a[1]
            na=a[2]
            nb=na+1
            fmax=fa+fa*percent/100
            fmin=fa-fa*percent/100
            self.cur.execute("SELECT candidate,frequency FROM term_candidates where frequency <="+str(fmax)+" and frequency>="+str(fmin)+"  and n ="+str(nb))
            results2=self.cur.fetchall()
            for b in results2:
                tb=b[0]
                fb=b[1]
                if not ta==tb and not tb.find(ta)==-1: 
                    self.cur.execute('DELETE FROM term_candidates WHERE candidate=?', (ta,))
                    if verbose:
                        print(str(fa),ta,"-->",str(fb),tb)
        self.conn.commit()

    def regexp_exclusion(self,verbose=False):
        '''Deletes term candidates matching a set of regular expresions loaded with the load_sl_exclusion_regexps method.'''
        self.cur.execute("SELECT sl_exclusion_regexp FROM sl_exclusion_regexps")
        results=self.cur.fetchall()
        for r in results:
            nregexp=len(r[0].split())
            exreg=r[0]
            self.cur.execute("SELECT candidate FROM term_candidates")
            results=self.cur.fetchall()
            cexreg=re.compile(exreg)
            for a in results:
                candidate=a[0]
                ncandidate=len(candidate.split())
                match=re.match(cexreg,candidate)
                if not match==None and nregexp==ncandidate:
                    self.cur.execute('DELETE FROM term_candidates WHERE candidate=?', (candidate,))
                    if verbose:
                        print(exreg,"-->",candidate)
            self.conn.commit()

    #EVALUATION
    
    
     
    def evaluate_pos(self,limit,order="desc",iterations=1000,ignore_case=True):
        '''Performs the evaluation of the term candidates using the evaluation_terms loaded with the load_evaluation_terms method.'''
        correct=0
        total=0
        evaluation_terms=[]
        self.cur.execute("SELECT sl_term FROM evaluation_terms")
        results=self.cur.fetchall()
        for r in results:
            evaluation_terms.append(r[0])
        tsr_terms=[]
        self.cur.execute("SELECT term FROM tsr_terms")
        results=self.cur.fetchall()
        for r in results:
            tsr_terms.append(r[0])
        evaluation_terms.extend(self.tsr_terms)
        with self.conn:
            for i in range(0,iterations):
                if order=="desc":
                    self.cur.execute("SELECT candidate,value from term_candidates where n<="+str(self.n_max)+" order by value desc, frequency desc, random() limit "+str(limit))
                elif order=="asc":
                    self.cur.execute("SELECT candidate from term_candidates where n<="+str(self.n_max)+" order by value asc, frequency desc, random() limit "+str(limit))
                else:
                    raise NameError('Order must be desc (decending) or asc (ascending). Defaulf value: desc')
                #self.cur.execute("SELECT candidate from term_candidates order by id limit "+str(limit))
                for s in self.cur.fetchall():
                    total+=1
                    candidate=s[0]
                    if ignore_case:
                        if candidate in evaluation_terms:
                            correct+=1
                        elif candidate.lower() in evaluation_terms:
                            correct+=1
                    else:
                        if candidate in evaluation_terms:
                            correct+=1
            correct=correct/iterations
            total=total/iterations
            
        try:
            precisio=100*correct/total
            recall=100*correct/len(evaluation_terms)
            f1=2*precisio*recall/(precisio+recall)
            return(limit,correct,total,precisio,recall,f1)
        except:
            return(limit,0,0,0,0,0)

    def association_measures(self,measure="raw_freq"):
        measurename=measure
        bigram_measures = myBigramAssocMeasures()
        trigram_measures = myTrigramAssocMeasures()
        quadgram_measures = myQuadgramAssocMeasures()
        
        fd_tokens=nltk.FreqDist()
        fd_bigrams=nltk.FreqDist()
        fd_trigrams=nltk.FreqDist()
        fd_quadgrams=nltk.FreqDist()
        wildcard_fd=nltk.FreqDist()
        self.cur.execute("SELECT token,frequency from tokens")
        for s in self.cur.fetchall():
            aux=(s[0])
            fd_tokens[aux]+=s[1]
            
        textcorpus=[]
        self.cur.execute("SELECT segment from sl_corpus")
        for segment in self.cur.fetchall():
            textcorpus.extend(segment[0].split())
            
        bigram_finder=BigramCollocationFinder.from_words(textcorpus)
        trigram_finder=TrigramCollocationFinder.from_words(textcorpus)
        quadgram_finder=QuadgramCollocationFinder.from_words(textcorpus)
        
        self.cur.execute("SELECT ngram,frequency,n from ngrams")
        results=self.cur.fetchall()
        for r in results:
            data=[]
            data.append(r[0])
            self.cur2.execute("UPDATE term_candidates SET value=NULL where candidate=?",data)
        self.conn.commit()
        data=[]
        bigram_measure=[]
        try:
            bigram_measure=eval("bigram_finder.score_ngrams(bigram_measures."+measure+")")
        except:
            print("WARNING: measure "+measure+ " not implemented for bigrams",sys.exc_info())
            #sys.exit()
            
        for nose in bigram_measure:
            record=[]
            term_candidate=" ".join(nose[0])
            mvalue=nose[1]
            record.append(measure)
            record.append(mvalue)
            record.append(term_candidate)
            data.append(record)
        
        trigram_measure=[]        
        try:
            trigram_measure=eval("trigram_finder.score_ngrams(trigram_measures."+measure+")")
        except:
            print("WARNING: measure "+measure+ " not implemented for trigrams")
            #sys.exit()
        for nose in trigram_measure:
            record=[]
            term_candidate=" ".join(nose[0])
            mvalue=nose[1]
            record.append(measure)
            record.append(mvalue)
            record.append(term_candidate)
            data.append(record)
        quadgram_measure=[]  
        try:
            quadgram_measure=eval("quadgram_finder.score_ngrams(quadgram_measures."+measure+")")
        except:
            print("WARNING: measure "+measure+ " not implemented for quadgrams")
            #sys.exit()
        
        for nose in quadgram_measure:
            record=[]
            term_candidate=" ".join(nose[0])
            mvalue=nose[1]
            record.append(measure)
            record.append(mvalue)
            record.append(term_candidate)
            data.append(record)
            
        self.conn.executemany("UPDATE term_candidates SET measure=?,value=? where candidate=?",data)
        self.conn.commit()
    


    def index_phrase_table(self,phrasetable):
        '''Indexes a phrase table from Moses.'''
        self.entrada=gzip.open(phrasetable, mode='rt',encoding='utf-8')

        self.pt={}
        self.continserts=0
        self.record=[]
        self.data=[]
        while 1:
            self.linia=self.entrada.readline()
            if not self.linia:
                break
            self.linia=self.linia.rstrip()
            self.camps=self.linia.split(" ||| ")
            self.source=self.camps[0].strip()
            self.trad=self.camps[1].strip()
            self.probs=self.camps[2].split()
            try:
                if not self.trad[0] in self.punctuation and not self.source[0] in self.punctuation and not self.trad[-1] in self.punctuation and not self.source[-1] in self.punctuation:
                    #Currently, four different phrase translation scores are computed:
                    #0    inverse phrase translation probability φ(f|e)
                    #1    inverse lexical weighting lex(f|e)
                    #2    direct phrase translation probability φ(e|f)
                    #3    direct lexical weighting lex(e|f)
                    #self.probtrad=float(self.probs[1])
                    self.probtrad=(float(self.probs[2])*float(self.probs[3]))
                    #print(self.source,self.trad,self.probtrad)
                    self.record=[]
                    self.record.append(self.source)
                    self.record.append(self.trad)
                    self.record.append(self.probtrad)
                    self.data.append(self.record)
                    self.continserts+=1
                    if self.continserts==self.maxinserts:
                        self.cur.executemany("INSERT INTO index_pt (source, target, probability) VALUES (?,?,?)",self.data)
                        self.data=[]
                        self.continserts=0
                        self.conn.commit()
            except:
                pass
        with self.conn:
            self.cur.executemany("INSERT INTO index_pt (source, target, probability) VALUES (?,?,?)",self.data)    
        self.conn.commit()
                
        
    def find_terms_in_parallel_corpus(self,SLterms,maxdec=1,maxinc=2,candidates=5,maxlines=-1):
        tofind=[]
        result={}
        if isinstance(SLterms, str):
            if os.path.exists(SLterms):
                entrada=codecs.open(SLterms)
                for linia in entrada:
                    linia=linia.rstrip()
                    tofind.append(linia)
                entrada.close()
            else:
                tofind.append(SLterms)
        elif isinstance(SLterms, list):
            tofind.extend(SLterms)
        tl_stopwords=[]
        with self.conn:
            self.cur.execute("SELECT tl_stopword FROM tl_stopwords")
            for s in self.cur.fetchall():
                tl_stopwords.append(s[0])
       
        for SLterm in tofind:
            fd=FreqDist()
            fd.clear()
            result[SLterm]={}
            if maxlines==-1:
                self.cur.execute("SELECT segmentTL FROM parallel_corpus where INSTR(segmentSL,\""+SLterm+"\")")
            else:
                self.cur.execute("SELECT segmentTL FROM parallel_corpus where INSTR(segmentSL,\""+SLterm+"\") limit "+str(maxlines))
            TLsegments=self.cur.fetchall()
            if self.specificSLtokenizer:
                termtok=self.SLtokenizer.tokenize(SLterm)
            else:
                termtok=SLterm
            nSLterm=len(termtok.split())
            nmin=nSLterm-maxdec
            if nmin<1: nmin=1
            nmax=nSLterm+maxinc
            for TLsegment in TLsegments:
                if self.specificTLtokenizer:
                    TLsegmenttok=self.TLtokenizer.tokenize(TLsegment[0]).split()
                else:
                    TLsegmenttok=TLsegment[0].split()
                for n in range(nmin,nmax+1):
                    ngs=ngrams(TLsegmenttok, n)
                    for ng in ngs:
                        include=True
                        if ng[0] in tl_stopwords: include=False
                        if len(ng)>1 and ng[1] in tl_stopwords: include=False
                        if include:
                            detokcandidate=" ".join(ng)
                            if self.specificTLtokenizer:
                                detokcandidate=self.TLtokenizer.detokenize(detokcandidate)
                            fd[detokcandidate]+=1
                            
            totalf=fd.N()            
            for mc in fd.most_common(candidates):
                result[SLterm][mc[0]]=mc[1]/totalf
        return(result)
    
    def compoundify_sl_corpus(self,term):
        term2=term.replace(" ","▁")
        self.cur.execute("SELECT id, segment FROM sl_corpus where INSTR(segment,\""+term+"\")")
        trobats=self.cur.fetchall()
        for trobat in trobats:
            ident=trobat[0]
            segment=trobat[1]
            segment2=segment.replace(term,term2)
            self.cur.execute("UPDATE sl_corpus SET segment=? where id=?",(segment2,ident))
        self.conn.commit()
    
    def compoundify_tl_corpus(self,term):
        term2=term.replace(" ","▁")
        self.cur.execute("SELECT id, segment FROM tl_corpus where INSTR(segment,\""+term+"\")")
        trobats=self.cur.fetchall()
        for trobat in trobats:
            ident=trobat[0]
            segment=trobat[1]
            segment2=segment.replace(term,term2)
            self.cur.execute("UPDATE tl_corpus SET segment=? where id=?",(segment2,ident))
        self.conn.commit()
    
    def compoundify_tl_corpus_mod(self,term):
        term2=term.replace(" ","▁")
        self.cur.execute("SELECT id, segment FROM tl_corpus where INSTR(segment,\""+term+"\")")
        trobats=self.cur.fetchall()
        data=[]
        for trobat in trobats:
            ident=trobat[0]
            segment=trobat[1]
            segment2=segment.replace(term,term2)
            data.append([segment2])
            
            
        self.cur.executemany("INSERT INTO tl_corpus (segment) VALUES (?)",data)
        self.conn.commit()
    
    def find_translation_comparable_corpus(self,SLterms,tl_stopwords=None,mapping_dictionary="MUSE-en-es.txt",maxdec=1,maxinc=2,candidates=25,compoundifySL=True,compoundifyTL=True,max_term_candidates_compoundify=200):
        tofind=[]
        result={}
        
        if isinstance(SLterms, str):
            if os.path.exists(SLterms):
                entrada=codecs.open(SLterms)
                for linia in entrada:
                    linia=linia.rstrip()
                    tofind.append(linia)
                entrada.close()
            else:
                tofind.append(SLterms)
        elif isinstance(SLterms, list):
            tofind.extend(SLterms)
        #compoundify SL corpus
        slnmin=1000000
        slnmax=0
        for SLterm in tofind:
            if self.specificSLtokenizer:
                termtok=self.SLtokenizer.tokenize(SLterm)
            else:
                termtok=SLterm
            if len(termtok.split())>1 and compoundifySL:
                self.compoundify_sl_corpus(SLterm)
            if len(termtok.split())<slnmin:slnmin=len(termtok.split())
            if len(termtok.split())>slnmax:slnmax=len(termtok.split())
            n_min=slnmin-maxdec
            if n_min<2: n_min=2
            n_max=slnmax+maxdec
        #compoundify TL corpus  (basic statistical term extraction)
        if compoundifyTL:
            self.delete_tokens()
            self.delete_ngrams()
            self.delete_sl_stopwords()
            self.delete_sl_inner_stopwords
            self.delete_sl_exclusion_regexps()
            self.delete_term_candidates()
            self.ngram_calculation (n_min,n_max,minfreq=2,corpus="tl_corpus")
            if not tl_stopwords==None:
                self.load_tl_stopwords(tl_stopwords)
            self.statistical_term_extraction(minfreq=2,corpus="tl_corpus")
            self.cur.execute("SELECT candidate FROM term_candidates ORDER BY frequency desc limit "+str(max_term_candidates_compoundify)+";")
            trobats=self.cur.fetchall()
            for trobat in trobats:
                term=trobat[0]
                self.compoundify_tl_corpus(term)
        print("CALCULATING EMBEDDINGS SL")
        self.calculate_embeddings_sl("embeddingsSL.temp",vector_size=300, window=5)
        print("CALCULATING EMBEDDINGS TL")
        self.calculate_embeddings_tl("embeddingsTL.temp",vector_size=300, window=5)
        print("MAPPING EMBEDDINGS")        
        self.mapEmbeddings("embeddingsSL.temp","embeddingsTL.temp","mappedSL.tmp","mappedTL.tmp",mapping_dictionary)        
        self.load_SL_embeddings("mappedSL.tmp")
        self.load_TL_embeddings("mappedTL.tmp")
        stopwords=[]
        with self.conn:
            self.cur.execute("SELECT tl_stopword FROM tl_stopwords")
            for s in self.cur.fetchall():
                stopwords.append(s[0])
        results={}
        for SLterm in tofind:
            if self.specificSLtokenizer:
                termtok=self.SLtokenizer.tokenize(SLterm)
            else:
                termtok=SLterm
            lenterm=len(termtok.split())
            lenmin=lenterm-maxdec
            lenmax=lenterm+maxinc
            results[SLterm]={}
            translations=self.find_translation_wv(SLterm,ncandidates=1000)
            cont=0
            for translation in translations:
                if self.specificTLtokenizer:
                    translationtok=self.TLtokenizer.tokenize(translation)
                else:
                    translationtok=translation
                lentranslation=len(translationtok.split())
                try:
                    if not translation in stopwords and not translation.split()[0] in stopwords and not translation.split()[-1] in stopwords and lentranslation>=lenmin and lentranslation<=lenmax:
                        results[SLterm][translation]=translations[translation]
                        cont+=1
                except:
                    pass
                if cont>=candidates:
                    break
    
        return(results)
    
    
    def find_translation_ptable(self,sourceterm,maxdec=1,maxinc=1,ncandidates=5,separator=":"):
        '''Finds translation equivalents in an indexed phrase table table. Requires an indexed phrase table and a a list of terms separated by ":".
        The number of translation candidates can be fixed, as well as the maximum decrement and increment of the number of tokens of the translation candidate'''
        #select target from index_pt where source="international conflict";
        self.cur.execute('SELECT target,probability FROM index_pt where source =?',(sourceterm,))
        self.results=self.cur.fetchall()
        self.targetcandidates={}
        for self.a in self.results:
            self.targetterm=self.a[0]
            self.probability=float(self.a[1])
            self.tttokens=self.targetterm.split()
            
            if not self.tttokens[0] in self.tl_stopwords and not self.tttokens[-1] in self.tl_stopwords and len(self.tttokens)>=len(sourceterm.split())-maxdec and len(self.tttokens)<=len(sourceterm.split())+maxinc:
                self.targetcandidates[self.targetterm]=self.probability
        self.sorted_x = sorted(self.targetcandidates.items(), key=operator.itemgetter(1),reverse=True)
        self.results=[]
        for self.s in self.sorted_x:
            self.results.append(self.s[0].replace(":",";"))
        return(separator.join(self.results[0:ncandidates]))
        
                
   
    def start_freeling_api(self,freelingpath, LANG):
        
        if not freelingpath.endswith("/"):freelingpath=freelingpath+"/"
        try:
            sys.path.append(freelingpath+"APIs/python3/")
            import pyfreeling
        except:
            print("No Freeling API available. Verify Freeling PATH: "+freelingpath+"freeling/APIs/python3/")
        
        pyfreeling.util_init_locale("default");

        # create language analyzer
        la1=pyfreeling.lang_ident(freelingpath+"common/lang_ident/ident.dat");

        # create options set for maco analyzer. Default values are Ok, except for data files.
        op1= pyfreeling.maco_options(LANG);
        op1.set_data_files( "", 
                           freelingpath + "common/punct.dat",
                           freelingpath+ LANG + "/dicc.src",
                           freelingpath + LANG + "/afixos.dat",
                           "",
                           freelingpath + LANG + "/locucions.dat", 
                           freelingpath + LANG + "/np.dat",
                           freelingpath + LANG + "/quantities.dat",
                           freelingpath + LANG + "/probabilitats.dat");

        # create analyzers
        self.tk1=pyfreeling.tokenizer(freelingpath+LANG+"/tokenizer.dat");
        self.sp1=pyfreeling.splitter(freelingpath+LANG+"/splitter.dat");
        self.sid1=self.sp1.open_session();
        self.mf1=pyfreeling.maco(op1);

        # activate mmorpho odules to be used in next call
        #(self, umap: "bool", num: "bool", pun: "bool", dat: "bool",
        # dic: "bool", aff: "bool", comp: "bool", rtk: "bool", 
        # mw: "bool", ner: "bool", qt: "bool", prb: "bool")
        #deactivate mw
        self.mf1.set_active_options(False, True, True, False,  # select which among created 
                              True, True, False, True,  # submodules are to be used. 
                              False, False, True, True ); # default: all created submodules are used

        # create tagger, sense anotator, and parsers
        self.tg1=pyfreeling.hmm_tagger(freelingpath+LANG+"/tagger.dat",True,2);
        
    def tag_freeling_api(self,corpus="source"):
        with self.conn:
            data=[]
            if corpus=="source":
                self.cur.execute('SELECT id,segment from sl_corpus')
            elif corpus=="target":
                self.cur.execute('SELECT id,segment from tl_corpus')
            continserts=0
            for s in self.cur.fetchall():
                id=s[0]
                segment=s[1]
                continserts+=1
                l1 = self.tk1.tokenize(segment);
                ls1 = self.sp1.split(self.sid1,l1,True);
                ls1 = self.mf1.analyze(ls1);
                ls1 = self.tg1.analyze(ls1);
                ttsentence=[]
                for s in ls1 :
                  ws = s.get_words();
                  for w in ws :
                    form=w.get_form()
                    lemma=w.get_lemma()
                    tag=w.get_tag()
                    ttsentence.append(form+"|"+lemma+"|"+tag)
                ttsentence=" ".join(ttsentence)
                record=[]
                record.append(id)
                record.append(ttsentence)
                data.append(record)
                if continserts==self.maxinserts:
                    if corpus=="source":
                        self.cur.executemany("INSERT INTO sl_tagged_corpus (id, tagged_segment) VALUES (?,?)",data)
                    if corpus=="target":
                        self.cur.executemany("INSERT INTO tl_tagged_corpus (id, tagged_segment) VALUES (?,?)",data)
                    data=[]
                    continserts=0
            with self.conn:
                if corpus=="source":
                    self.cur.executemany("INSERT INTO sl_tagged_corpus (id, tagged_segment) VALUES (?,?)",data) 
                if corpus=="target":
                    self.cur.executemany("INSERT INTO tl_tagged_corpus (id, tagged_segment) VALUES (?,?)",data)
    
    
    #SPACY TAGGER
    def load_POS_model_spacy(self, model):
        if not spacy.util.is_package(model):
            print("Downloading and installing ",model)
            try:
                subprocess.check_call([sys.executable, "-m", "spacy", "download", model])
                print("Model downloaded. Stopping the program. The program should be run again to load the downloaded model.")
            except:
                print("Model",model,"not available.")
        else:
            self.POSmodel_spacy=spacy.load(model)
        
    def tag_spacy(self,corpus="source",mode="coarse"):
        #mode on of coarse or fine
        with self.conn:
            data=[]
            if corpus=="source":
                self.cur.execute('SELECT id,segment from sl_corpus')
            elif corpus=="target":
                self.cur.execute('SELECT id,segment from tl_corpus')
            elif corpus=="parallel-source":
                self.cur.execute('SELECT id,segmentSL from parallel_corpus')
            elif corpus=="parallel-target":
                self.cur.execute('SELECT id,segmentTL from parallel_corpus')
            continserts=0
            for s in self.cur.fetchall():
                id=s[0]
                segment=s[1]
                continserts+=1
                taggedtokens = self.POSmodel_spacy(segment)
                ttsentence=[]
                for token in taggedtokens:
                    form=token.text
                    lemma=token.lemma_
                    if mode=="fine":
                        tag=token.tag_
                    elif mode=="coarse":
                        tag=token.pos_    
                    ttsentence.append(form+"|"+lemma+"|"+tag)
                ttsentence=" ".join(ttsentence)
                record=[]
                record.append(id)
                record.append(ttsentence)
                data.append(record)
                if continserts==self.maxinserts:
                    if corpus=="source":
                        self.cur.executemany("INSERT INTO sl_tagged_corpus (id, tagged_segment) VALUES (?,?)",data)
                    elif corpus=="target":
                        self.cur.executemany("INSERT INTO tl_tagged_corpus (id, tagged_segment) VALUES (?,?)",data)
                    elif corpus=="parallel-source":
                        self.cur.executemany("INSERT INTO tagged_parallel_corpus (id, tagged_segmentSL) VALUES (?,?) ON CONFLICT (id) DO UPDATE  SET tagged_segmentSL=excluded.tagged_segmentSL",data)
                    elif corpus=="parallel-target":
                        self.cur.executemany("INSERT INTO tagged_parallel_corpus (id, tagged_segmentTL) VALUES (?,?) ON CONFLICT (id) DO UPDATE SET tagged_segmentTL=excluded.tagged_segmentTL",data)
                    data=[]
                    continserts=0
            with self.conn:
                if corpus=="source":
                    self.cur.executemany("INSERT INTO sl_tagged_corpus (id, tagged_segment) VALUES (?,?)",data) 
                elif corpus=="target":
                    self.cur.executemany("INSERT INTO tl_tagged_corpus (id, tagged_segment) VALUES (?,?)",data)
                elif corpus=="parallel-source":
                    self.cur.executemany("INSERT INTO tagged_parallel_corpus (id, tagged_segmentSL) VALUES (?,?) ON CONFLICT (id) DO UPDATE  SET tagged_segmentSL=excluded.tagged_segmentSL",data)
                elif corpus=="parallel-target":
                    self.cur.executemany("INSERT INTO tagged_parallel_corpus (id, tagged_segmentTL) VALUES (?,?) ON CONFLICT (id) DO UPDATE SET tagged_segmentTL=excluded.tagged_segmentTL",data)
    
    #SPACY_UDPIPE TAGGER
    def load_POS_model_spacy_udpipe(self, language):
        try:
            self.POSmodel = spacy_udpipe.load(language)
        except:
            print("No model for ",language," available.")
            print("Downloading and installing model for ",language)
            try:
                spacy_udpipe.download(language)
                self.POSmodel = spacy_udpipe.load(language)
            except:
                print("ERROR: not able to load spacy_udepipe model for ",language)
            
        
    def tag_spacy_udpipe(self,corpus="source"):
        #mode on of coarse or fine
        with self.conn:
            data=[]
            if corpus=="source":
                self.cur.execute('SELECT id,segment from sl_corpus')
            elif corpus=="target":
                self.cur.execute('SELECT id,segment from tl_corpus')
            elif corpus=="parallel-source":
                self.cur.execute('SELECT id,segmentSL from parallel_corpus')
            elif corpus=="parallel-target":
                self.cur.execute('SELECT id,segmentTL from parallel_corpus')
            continserts=0
            for s in self.cur.fetchall():
                id=s[0]
                segment=s[1]
                continserts+=1
                taggedtokens = self.POSmodel(segment)
                ttsentence=[]
                for token in taggedtokens:
                    form=token.text
                    lemma=token.lemma_
                    tag=token.tag_
                    tag=token.pos_    
                    ttsentence.append(form+"|"+lemma+"|"+tag)
                ttsentence=" ".join(ttsentence)
                record=[]
                record.append(id)
                record.append(ttsentence)
                data.append(record)
                if continserts==self.maxinserts:
                    if corpus=="source":
                        self.cur.executemany("INSERT INTO sl_tagged_corpus (id, tagged_segment) VALUES (?,?)",data)
                    elif corpus=="target":
                        self.cur.executemany("INSERT INTO tl_tagged_corpus (id, tagged_segment) VALUES (?,?)",data)
                    elif corpus=="parallel-source":
                        self.cur.executemany("INSERT INTO tagged_parallel_corpus (id, tagged_segmentSL) VALUES (?,?) ON CONFLICT (id) DO UPDATE  SET tagged_segmentSL=excluded.tagged_segmentSL",data)
                    elif corpus=="parallel-target":
                        self.cur.executemany("INSERT INTO tagged_parallel_corpus (id, tagged_segmentTL) VALUES (?,?) ON CONFLICT (id) DO UPDATE SET tagged_segmentTL=excluded.tagged_segmentTL",data)
                    data=[]
                    continserts=0
            with self.conn:
                if corpus=="source":
                    self.cur.executemany("INSERT sl_tagged_corpus (id, tagged_segment) VALUES (?,?)",data) 
                elif corpus=="target":
                    self.cur.executemany("INSERT INTO tl_tagged_corpus (id, tagged_segment) VALUES (?,?)",data)
                elif corpus=="parallel-source":
                    self.cur.executemany("INSERT INTO tagged_parallel_corpus (id, tagged_segmentSL) VALUES (?,?) ON CONFLICT (id) DO UPDATE  SET tagged_segmentSL=excluded.tagged_segmentSL",data)
                elif corpus=="parallel-target":
                    self.cur.executemany("INSERT INTO tagged_parallel_corpus (id, tagged_segmentTL) VALUES (?,?) ON CONFLICT (id) DO UPDATE SET tagged_segmentTL=excluded.tagged_segmentTL",data)
    
    
    def save_sl_tagged_corpus(self,outputfile,encoding="utf-8"):
        sortida=codecs.open(outputfile,"w",encoding=encoding)
        self.cur.execute('SELECT tagged_segment from sl_tagged_corpus')
        for s in self.cur.fetchall():
            tagged_segment=s[0]
            sortida.write(tagged_segment+"\n")
    
    def save_tl_tagged_corpus(self,outputfile,encoding="utf-8"):
        sortida=codecs.open(outputfile,"w",encoding=encoding)
        self.cur.execute('SELECT tagged_segment from tl_tagged_corpus')
        for s in self.cur.fetchall():
            tagged_segment=s[0]
            sortida.write(tagged_segment+"\n")
            
            
            
    def save_sl_tagged_parallel_corpus(self,outputfile,encoding="utf-8"):
        sortida=codecs.open(outputfile,"w",encoding=encoding)
        self.cur.execute('SELECT tagged_segmentSL from tagged_parallel_corpus')
        for s in self.cur.fetchall():
            tagged_segment=s[0]
            sortida.write(tagged_segment+"\n")
            
    def save_tl_tagged_parallel_corpus(self,outputfile,encoding="utf-8"):
        sortida=codecs.open(outputfile,"w",encoding=encoding)
        self.cur.execute('SELECT tagged_segmentTL from tagged_parallel_corpus')
        for s in self.cur.fetchall():
            tagged_segment=s[0]
            sortida.write(tagged_segment+"\n")

    
    def tagged_ngram_calculation (self,nmin=2,nmax=3,minfreq=2):
        '''Calculates the tagged ngrams.'''
        ngramsFD=FreqDist()
        n_nmin=nmin
        n_max=nmax
        data=[]
        record=[]
        with self.conn:
            self.cur.execute('SELECT tagged_segment from sl_tagged_corpus')
            for s in self.cur.fetchall():
                segment=s[0]
                for n in range(nmin,nmax+1):
                    ngs=ngrams(segment.split(),n)
                    for ng in ngs:
                        ngramsFD[ng]+=1
        for c in ngramsFD.most_common():
           if c[1]>=minfreq:
                candidate=[]
                for ngt in c[0]:
                    candidate.append(ngt.split("|")[0])
                candidate=" ".join(candidate)
                record=[]
                record.append(candidate)
                record.append(" ".join(c[0])) 
                record.append(len(c[0]))
                record.append(c[1])   
                data.append(record)
        with self.conn:
            self.cur.executemany("INSERT INTO tagged_ngrams (ngram, tagged_ngram, n, frequency) VALUES (?,?,?,?)",data) 
            self.conn.commit()
                    
    def translate_linguistic_pattern(self,pattern):
        aux=[]
        for ptoken in pattern.split():
            auxtoken=[]
            ptoken=ptoken.replace(".*","[^\s]+") 
            for pelement in ptoken.split("|"):
                if pelement=="#":
                    auxtoken.append("([^\s]+?)")                    
                elif pelement=="":
                    auxtoken.append("[^\s]+?")
                else:
                    if pelement.startswith("#"):
                        auxtoken.append("("+pelement.replace("#","")+")")
                    else:
                        auxtoken.append(pelement)
            aux.append("\|".join(auxtoken))
        tp="("+" ".join(aux)+")"
        return(tp)       
    
    def load_linguistic_patterns(self,file, encoding="utf-8"):
        '''Loads the linguistic patterns to use with linguistic terminology extraction.'''
        entrada=codecs.open(file,"r",encoding=encoding)
        linguistic_patterns=[]
        data=[]
        record=[]
        for linia in entrada:
            linia=linia.rstrip()
            npattern=len(linia.split(" "))
            if npattern<self.n_min_pos_patterns: self.n_min_pos_patterns=npattern
            if npattern>self.n_max_pos_patterns: self.n_max_pos_patterns=npattern
            pattern=self.translate_linguistic_pattern(linia)
            record.append(pattern)
            data.append(record)
            record=[]
        with self.conn:
            self.cur.executemany("INSERT INTO linguistic_patterns (linguistic_pattern) VALUES (?)",data)
    def get_n_min_pos_patterns(self):
        return(self.n_min_pos_patterns)
        
    def get_n_max_pos_patterns(self):
        return(self.n_max_pos_patterns)
    
    def linguistic_term_extraction(self,minfreq=2):
        '''Performs an linguistic term extraction using the extracted tagged ngrams (tagged_ngram_calculation should be executed first). '''
        linguistic_patterns=[]
        controlpatterns=[]
        with self.conn:
            self.cur.execute("SELECT linguistic_pattern from linguistic_patterns")
            for lp in self.cur.fetchall():
                linguistic_pattern=lp[0]
                transformedpattern="^"+linguistic_pattern+"$"
                if not transformedpattern in controlpatterns:
                    linguistic_patterns.append(transformedpattern)
                    controlpatterns.append(transformedpattern)            
        self.cur.execute("SELECT tagged_ngram, n, frequency FROM tagged_ngrams order by frequency desc")
        results=self.cur.fetchall()
        data=[] 
        for a in results:
            include=True
            ng=a[0]
            n=a[1]
            frequency=a[2]
            try:
                if ng.split()[0].split("|")[1].lower() in sl_stopwords: include=False
            except:
                pass
            try:
                if ng.split()[-1].split("|")[1].lower() in sl_stopwords: include=False
            except:
                pass
            if frequency<minfreq:
                break
            if include:
                for pattern in linguistic_patterns:
                    match=re.search(pattern,ng)
                    if match:
                        if match.group(0)==ng:          
                            candidate=" ".join(match.groups()[1:])
                            record=[]
                            record.append(candidate)     
                            record.append(n)
                            record.append(frequency)   
                            record.append("freq")
                            record.append(frequency)   
                            data.append(record)
                            break
        with self.conn:
            self.cur.executemany("INSERT INTO term_candidates (candidate, n, frequency, measure, value) VALUES (?,?,?,?,?)",data)      
            self.conn.commit()
            
        #deleting repeated candidates
        self.cur.execute("SELECT candidate, n, frequency FROM term_candidates")
        results=self.cur.fetchall()
        tcaux={}
        for a in results:
            if not a[0] in tcaux:
                tcaux[a[0]]=a[2]
            else:
                tcaux[a[0]]+=a[2]
        self.cur.execute("DELETE FROM term_candidates")
        self.conn.commit()
        data=[] 
        for tc in tcaux:
            record=[]
            record.append(tc)            
            record.append(len(tc.split()))
            record.append(tcaux[tc])   
            record.append("freq")
            record.append(tcaux[tc])   
            data.append(record)
        with self.conn:
            self.cur.executemany("INSERT INTO term_candidates (candidate, n, frequency, measure, value) VALUES (?,?,?,?,?)",data) 
            self.conn.commit()
            
    def learn_linguistic_patterns(self,outputfile,showfrequencies=False,encoding="utf-8",verbose=True,representativity=100):
        learntpatterns={}
        sortida=codecs.open(outputfile,"w",encoding=encoding)
        acufreq=0
        tags={}
        with self.conn:
            self.cur.execute("SELECT sl_term FROM evaluation_terms")
            for s in self.cur.fetchall():
                self.cur.execute("SELECT tagged_ngram, n, frequency FROM tagged_ngrams WHERE ngram= ?", (s[0],))
                results=self.cur.fetchall()
                if len(results)>0:
                    for a in results:
                        ng=a[0]
                        nglist=ng.split()
                        n=a[1]
                        frequency=a[2]
                        candidate=[]
                        ngtokenstag=ng.split()
                        for ngt in ngtokenstag:
                            candidate.append(ngt.split("|")[0])
                        candidate=" ".join(candidate)
                        t2=ng.split()
                        t1=candidate.split()
                        patternbrut=[]
                        for position in range(0,n):
                            t2f=t2[position].split("|")[0]
                            t2l=t2[position].split("|")[1]
                            t2t=t2[position].split("|")[2]
                            patternpart=""
                            if t1[position]==t2l:
                                patternpart="|#|"+t2t
                            elif t1[position]==t2f:
                                patternpart="#||"+t2t
                            patternbrut.append(patternpart)
                        pattern=" ".join(patternbrut)
                        if pattern in learntpatterns:
                            learntpatterns[pattern]+=n
                            acufreq+=n
                        else:
                            learntpatterns[pattern]=n
                            acufreq+=n
        sorted_x = sorted(learntpatterns.items(), key=operator.itemgetter(1),reverse=True)
        results=[]
        acufreq2=0
        for s in sorted_x:
            percent=100*acufreq2/acufreq
            if percent>representativity:
                break
            acufreq2+=s[1]
            if showfrequencies:
                cadena=str(s[1])+"\t"+s[0]
            else:
                cadena=s[0]
            sortida.write(cadena+"\n")
            if verbose:
                print(cadena)
                                
    def find_translation_pcorpus_statistical(self,slterm,maxdec=1,maxinc=1,ncandidates=5,separator=":"):
        self.nmin=len(slterm.split())-maxdec
        self.nmax=len(slterm.split())+maxinc
        self.tlngrams=FreqDist()
        with self.conn:
            self.cur.execute('SELECT id, segment from sl_corpus')
            
            for self.s in self.cur.fetchall():
                self.segment=self.s[1]
                self.id=self.s[0]
                
                if self.segment.find(slterm)>-1:
                    self.cur2.execute('SELECT segment from tl_corpus where id="'+str(self.id)+'"')
                    for self.s2 in self.cur2.fetchall():
                        self.tl_segment=self.s2[0]
                        for self.n in range(self.nmin,self.nmax+1):
                            #self.tlngs=ngrams(self.tl_tokenizer.tokenize(self.tl_segment), self.n)
                            self.tlngs=ngrams(self.tl_segment.split(), self.n)
                            for self.tlng in self.tlngs:
                                if not self.tlng[0] in self.tl_stopwords and not self.tlng[-1] in self.tl_stopwords:
                                    self.tlngrams[self.tlng]+=1
            self.resultlist=[]
            for self.c in self.tlngrams.most_common(ncandidates):
                self.resultlist.append(" ".join(self.c[0]))
            
            return(separator.join(self.resultlist))
    
    def find_translation_pcorpus_linguistics(self,slterm,maxdec=1,maxinc=1,ncandidates=5,separator=":"):
        self.nmin=len(slterm.split())-maxdec
        self.nmax=len(slterm.split())+maxinc
        self.tlngrams=FreqDist()
        with self.conn:
            self.cur.execute('SELECT id, segment from sl_corpus')
            
            for self.s in self.cur.fetchall():
                self.segment=self.s[1]
                self.id=self.s[0]
                
                if self.segment.find(slterm)>-1:
                    self.cur2.execute('SELECT segment from tl_corpus where id="'+str(self.id)+'"')
                    for self.s2 in self.cur2.fetchall():
                        self.tl_segment=self.s2[0]
                        for self.n in range(self.nmin,self.nmax+1):
                            #self.tlngs=ngrams(self.tl_tokenizer.tokenize(self.tl_segment), self.n)
                            self.tlngs=ngrams(self.tl_segment.split(), self.n)
                            for self.tlng in self.tlngs:
                                if not self.tlng[0] in self.tl_stopwords and not self.tlng[-1] in self.tl_stopwords:
                                    self.tlngrams[self.tlng]+=1
            self.resultlist=[]
            for self.c in self.tlngrams.most_common(ncandidates):
                self.resultlist.append(" ".join(self.c[0]))
            
            return(separator.join(self.resultlist))

#EMBEDDINGS

    def calculate_embeddings_sl(self,filename,vector_size=300, window=5, min_count=1, workers=4):
        self.cur.execute('SELECT id, segment from sl_corpus')
        data = []
        for s in self.cur.fetchall():
            temp=[]
            segment=s[1] 
            if self.specificSLtokenizer:
                tokens=self.SLtokenizer.tokenize(segment).split()
            else:
                tokens=segment.split()
            data.append(tokens)
        model = Word2Vec(sentences=data, vector_size=vector_size, window=window, min_count=min_count, workers=workers)
        model.wv.save_word2vec_format(filename, binary=False)
     
    def calculate_embeddings_sl_ref(self,filename,vector_size=300, window=5, min_count=1, workers=4):
        self.cur.execute('SELECT id, segment from tl_corpus')
        data = []
        for s in self.cur.fetchall():
            temp=[]
            segment=s[1] 
            if self.specificSLtokenizer:
                tokens=self.SLtokenizer.tokenize(segment).split()                
            else:
                tokens=segment.split()
            data.append(tokens)
        model = Word2Vec(sentences=data, vector_size=vector_size, window=window, min_count=min_count, workers=workers)
        model.wv.save_word2vec_format(filename, binary=False)
    
    def calculate_embeddings_tl(self,filename,vector_size=300, window=5, min_count=1, workers=4):
        self.cur.execute('SELECT id, segment from tl_corpus')
        data = []
        for s in self.cur.fetchall():
            temp=[]
            segment=s[1] 
            if self.specificTLtokenizer:
                tokens=self.TLtokenizer.tokenize(segment).split()
            else:
                tokens=segment.split()
            data.append(tokens)
        model = Word2Vec(sentences=data, vector_size=vector_size, window=window, min_count=min_count, workers=workers)
        model.wv.save_word2vec_format(filename, binary=False)
    
    def mapEmbeddings(self,src_input,trg_input,src_output,trg_output,init_dictionary):
        supervised_mapping(src_input,trg_input,src_output,trg_output,init_dictionary)
        
    def load_SL_embeddings(self, file, binary=False):
        self.wvSL = KeyedVectors.load_word2vec_format(file, binary=False)
        
    def load_TL_embeddings(self, file, binary=False):
        self.wvTL = KeyedVectors.load_word2vec_format(file, binary=False)
        
    
    def find_translation_wv(self, term, ncandidates=50):
        
        term=term.strip().replace(" ","▁")
        try:
            vector=self.wvSL[term]
            tcandidates = self.wvTL.most_similar([vector], topn=ncandidates)
        except:
            tcandidates=[]
        response={}
        
        for tc in tcandidates:
            tc2=tc[0].replace("▁"," ")
            response[tc2]=tc[1]
                
        return(response)
        


#TSR
    def tsr(self, type="combined",max_iterations=10000000000, verbose=True): 
        component={}
        firstcomponent={}
        middlecomponent={}
        lastcomponent={}
        self.tsr_terms=[]
        self.cur.execute("SELECT term FROM tsr_terms")
        results=self.cur.fetchall()
        for r in results:
            self.tsr_terms.append(r[0])
        for term in self.tsr_terms:
            camps=term.split()
            if len(camps)==1: #UNIGRAMS
                firstcomponent[camps[0].lower()]=1
                lastcomponent[camps[0].lower()]=1
            if len(camps)>=2:
                firstcomponent[camps[0].lower()]=1
                lastcomponent[camps[-1].lower()]=1
                component[camps[0].lower()]=1
                component[camps[-1].lower()]=1
                if len(camps)>=3:
                    for i in range(1,len(camps)-1):
                        middlecomponent[camps[i].lower()]=1
                        component[camps[i].lower()]=1

        new=True
        newcandidates={} #candidate-frequency
        hashmeasure={}
        hashvalue={}
        
        newcandidatestempstric={} #candidate-frequency                                                                    
        hashmeasuretempstrict={}
        hashvaluetempstric={}

        newcandidatestempflexible={} #candidate-frequency
        hashmeasuretempflexible={}
        hashvaluetempflexible={}
        
        newcandidatestempcombined={} #candidate-frequency
        hashmeasuretempcombined={}
        hashvaluetempcombined={}
        
        iterations=0
        while new:
            iterations+=1
            if verbose: print("ITERATION",iterations)
            new=False
            self.cur.execute("SELECT candidate,n,frequency,measure,value FROM term_candidates ")
            results=self.cur.fetchall()
            auxiliar={}
            value=max_iterations-iterations#r[4]
            for r in results:
                candidate=r[0]
                n=r[1]
                frequency=r[2]
                measure="tsr"#r[3]
                #IMPLEMENTED ONLY FOR BIGRAMS !!!
                '''
                rcamps=candidate.split()
                if type=="strict":
                    if rcamps[0] in firstcomponent and rcamps[-1] in lastcomponent:
                        if not candidate in newcandidates:
                            newcandidates[candidate]=frequency
                            hashmeasure[candidate]=measure
                            hashvalue[candidate]=value
                            new=True
                            firstcomponent[rcamps[0]]=1
                            lastcomponent[rcamps[-1]]=1
                elif type=="flexible":
                    if rcamps[0] in firstcomponent or rcamps[-1] in lastcomponent:
                        if not candidate in newcandidates:
                            newcandidates[candidate]=frequency
                            hashmeasure[candidate]=measure
                            hashvalue[candidate]=value
                            new=True
                            firstcomponent[rcamps[0]]=1
                            lastcomponent[rcamps[-1]]=1
                            component[rcamps[0]]=1
                            component[rcamps[-1]]=1
                elif type=="combined":
                    if iterations==1:
                        if rcamps[0] in firstcomponent and rcamps[-1] in lastcomponent:
                            if not candidate in newcandidates:
                                newcandidates[candidate]=frequency
                                hashmeasure[candidate]=measure
                                hashvalue[candidate]=value
                                new=True
                                firstcomponent[rcamps[0]]=1
                                lastcomponent[rcamps[-1]]=1
                                component[rcamps[0]]=1
                                component[rcamps[-1]]=1
                    else:
                        if rcamps[0] in firstcomponent or rcamps[-1] in lastcomponent:
                            if not candidate in newcandidates:
                                newcandidates[candidate]=frequency
                                hashmeasure[candidate]=measure
                                hashvalue[candidate]=value
                                new=True
                                firstcomponent[rcamps[0]]=1
                                lastcomponent[rcamps[-1]]=1
                                component[rcamps[0]]=1
                                component[rcamps[-1]]=1
                '''
                first_c=False
                middle_c=False
                last_c=False
                rcamps=candidate.split()
                truesfalses=[]
                if str(rcamps[0]).lower() in firstcomponent: 
                    first_c=True
                    truesfalses.append(True)
                else:
                    truesfalses.append(False)
                if str(rcamps[-1]).lower() in lastcomponent: 
                    last_c=True
                    truesfalses.append(True)
                else:
                    truesfalses.append(False)
                if n>2:
                    middle_c=True
                    for i in range(1,n-1):
                        if not str(r[i]).lower() in middlecomponent: middle_c=False
                    if middle_c==True:
                        truesfalses.append(True)
                    else:
                        truesfalses.append(False)
                if type=="strict":
                    if not False in truesfalses:
                        if not candidate in newcandidates:
                            newcandidates[candidate]=frequency
                            hashmeasure[candidate]=measure
                            hashvalue[candidate]=value
                            new=True
                            firstcomponent[rcamps[0]]=1
                            lastcomponent[rcamps[-1]]=1
                elif type=="flexible":
                    if True in truesfalses:
                        if not candidate in newcandidates:
                            newcandidates[candidate]=frequency
                            hashmeasure[candidate]=measure
                            hashvalue[candidate]=value
                            new=True
                            firstcomponent[rcamps[0]]=1
                            lastcomponent[rcamps[-1]]=1
                            component[rcamps[0]]=1
                            component[rcamps[-1]]=1
                elif type=="combined":
                    if iterations==1:       
                        new=True
                        if not False in truesfalses:
                            if not candidate in newcandidates:
                                newcandidates[candidate]=frequency
                                hashmeasure[candidate]=measure
                                hashvalue[candidate]=value                                
                                firstcomponent[rcamps[0]]=1
                                lastcomponent[rcamps[-1]]=1
                                if n>2:
                                    for i in range(1,n-1):
                                        middlecomponent[rcamps[i]]=1
                                        component[rcamps[i]]=1
                    else:
                        if True in truesfalses:
                            if not candidate in newcandidates:
                                newcandidates[candidate]=frequency
                                hashmeasure[candidate]=measure
                                hashvalue[candidate]=value
                                new=True
                                firstcomponent[rcamps[0]]=1
                                lastcomponent[rcamps[-1]]=1
                                if n>2:
                                    for i in range(1,n-1):
                                        middlecomponent[rcamps[i]]=1
                                        component[rcamps[i]]=1
                                component[rcamps[0]]=1
                                component[rcamps[-1]]=1
                '''
                if n==2:
                    if rcamps[0] in firstcomponent: first_c=True
                    middle_c=True
                    if rcamps[-1] in lastcomponent: last_c=True
                    if type=="strict":
                        if first_c and last_c:
                            if not candidate in newcandidates:
                                newcandidates[candidate]=frequency
                                hashmeasure[candidate]=measure
                                hashvalue[candidate]=value
                                new=True
                                firstcomponent[rcamps[0]]=1
                                lastcomponent[rcamps[-1]]=1
                    elif type=="flexible":
                        if first_c or last_c:
                            if not candidate in newcandidates:
                                newcandidates[candidate]=frequency
                                hashmeasure[candidate]=measure
                                hashvalue[candidate]=value
                                new=True
                                firstcomponent[rcamps[0]]=1
                                lastcomponent[rcamps[-1]]=1
                                component[rcamps[0]]=1
                                component[rcamps[-1]]=1
                    
                    elif type=="combined":
                        if iterations==1:
                            if first_c and last_c:
                                if not candidate in newcandidates:
                                    newcandidates[candidate]=frequency
                                    hashmeasure[candidate]=measure
                                    hashvalue[candidate]=value
                                    new=True
                                    firstcomponent[rcamps[0]]=1
                                    lastcomponent[rcamps[-1]]=1
                                    component[rcamps[0]]=1
                                    component[rcamps[-1]]=1
                        else:
                            if first_c or last_c:
                                if not candidate in newcandidates:
                                    newcandidates[candidate]=frequency
                                    hashmeasure[candidate]=measure
                                    hashvalue[candidate]=value
                                    new=True
                                    firstcomponent[rcamps[0]]=1
                                    lastcomponent[rcamps[-1]]=1
                                    component[rcamps[0]]=1
                                    component[rcamps[-1]]=1
                                    
                elif n==3:
                    if rcamps[0] in firstcomponent: first_c=True
                    if rcamps[1] in middlecomponent: middle_c=True
                    if rcamps[-1] in lastcomponent: last_c=True
                    if type=="strict":
                        if first_c and middle_c and last_c:
                            if not candidate in newcandidates:
                                newcandidates[candidate]=frequency
                                hashmeasure[candidate]=measure
                                hashvalue[candidate]=value
                                new=True
                                firstcomponent[rcamps[0]]=1
                                middlecomponent[rcamps[1]]=1
                                lastcomponent[rcamps[-1]]=1
                    elif type=="flexible":
                        condition=False
                        if first_c and middle_c or last_c: condition=True
                        #if first_c or middle_c and last_c: condition=True
                        if last_c and middle_c or first_c: condition=True
                        if condition:
                            if not candidate in newcandidates:
                                newcandidates[candidate]=frequency
                                hashmeasure[candidate]=measure
                                hashvalue[candidate]=value
                                new=True
                                firstcomponent[rcamps[0]]=1
                                middlecomponent[rcamps[1]]=1
                                lastcomponent[rcamps[-1]]=1
                                component[rcamps[0]]=1
                                component[rcamps[1]]=1
                                component[rcamps[-1]]=1
                    
                    elif type=="combined":
                        if iterations==1:
                            if first_c and middle_c and last_c:
                                if not candidate in newcandidates:
                                    newcandidates[candidate]=frequency
                                    hashmeasure[candidate]=measure
                                    hashvalue[candidate]=value
                                    new=True
                                    firstcomponent[rcamps[0]]=1
                                    middlecomponent[rcamps[1]]=1
                                    lastcomponent[rcamps[-1]]=1
                                    component[rcamps[0]]=1
                                    component[rcamps[1]]=1
                                    component[rcamps[-1]]=1
                        else:
                            condition=False
                            if first_c or middle_c or last_c: condition=True
                            #if first_c and middle_c or last_c: condition=True
                            #if first_c or middle_c and last_c: condition=True
                            #if last_c and middle_c or first_c: condition=True
                            if condition:
                                if not candidate in newcandidates:
                                    newcandidates[candidate]=frequency
                                    hashmeasure[candidate]=measure
                                    hashvalue[candidate]=value
                                    new=True
                                    firstcomponent[rcamps[0]]=1
                                    middlecomponent[rcamps[1]]=1
                                    lastcomponent[rcamps[-1]]=1
                                    component[rcamps[0]]=1
                                    component[rcamps[1]]=1
                                    component[rcamps[-1]]=1
            '''     
            if iterations>=max_iterations:
                break
            if verbose: print(iterations,new)
        with self.conn:
            self.cur.execute('DELETE FROM term_candidates')
            self.conn.commit()
        
                    
        data=[]
        for c in newcandidates:
            termb=c
            n=len(c.split())
            freqtotal=newcandidates[c]
            measure=hashmeasure[c]
            value=hashvalue[c]
            record=[]
            record.append(termb)
            record.append(n)
            record.append(freqtotal)
            record.append(measure)
            record.append(value)
            data.append(record)
        with self.conn:
            self.cur.executemany("INSERT INTO term_candidates (candidate, n, frequency, measure, value) VALUES (?,?,?,?,?)",data)
            
        self.conn.commit()   

def L_LLR(a,b,c):
    '''Auxiliar function to calculate Log Likelihood Ratio'''
    L=(c**a)*((1-c)**(b-a))
    return(L)
            
class myBigramAssocMeasures(nltk.collocations.BigramAssocMeasures):
    """
    A collection of bigram association measures. Each association measure
    is provided as a function with three arguments::

        bigram_score_fn(n_ii, (n_ix, n_xi), n_xx)

    The arguments constitute the marginals of a contingency table, counting
    the occurrences of particular events in a corpus. The letter i in the
    suffix refers to the appearance of the word in question, while x indicates
    the appearance of any word. Thus, for example:

        n_ii counts (w1, w2), i.e. the bigram being scored
        n_ix counts (w1, *)
        n_xi counts (*, w2)
        n_xx counts (*, *), i.e. any bigram

    This may be shown with respect to a contingency table::

                w1    ~w1
             ------ ------
         w2 | n_ii | n_oi | = n_xi
             ------ ------
        ~w2 | n_io | n_oo |
             ------ ------
             = n_ix        TOTAL = n_xx
             
    Amb la terminologia de Pazienza

                w1    ~w1
             ------ ------
         w2 | n_ii O11| n_oi O12| = n_xi
             ------ ------
        ~w2 | n_io O21| n_oo O22|
             ------ ------
             = n_ix        TOTAL = n_xx
             
             N=O11+O12+O21+O22= n_xx
             R1=O11+O12 = n_xi
             R2=O21+O22
             C1=O11+O21=n_ix
             C2=O12+O22= n_oi+n_oo
             
        TENIM: n_ii, n_ix_xi_tuple, n_xx
        n_io=n_ix-n_ii
        n_oi=n_xi-n_ii
        n_oo=n_xx-n_ix-n_xi
    """
    #MEASURES NOT IMPLEMENTED IN NLTK
    
    def loglikelihood(self,n_ii, n_ix_xi_tuple, n_xx):
        '''LogLikelihood according to NSP'''
        (n_ix, n_xi) = n_ix_xi_tuple
        n_oo=n_xx-n_ix-n_xi
        n_io=n_ix-n_ii
        n_oi=n_xi-n_ii
        n_oo=n_xx-n_ix-n_xi        
        
        n11=n_ii
        n12=n_io
        n21=n_oi
        n22=n_oo
        n1p=n11+n12
        np1=n11+n21
        n2p=n21+n22
        np2=n12+n22
        npp=n_xx

        m11 = (n1p*np1/npp)
        m12 = (n1p*np2/npp)
        m21 = (np1*n2p/npp)
        m22 = (n2p*np2/npp)
        try:
            LogLikelihood = 2 * (n11 * math.log((n11/m11),2) + n12 * math.log((n12/m12),2) + n21 * math.log((n21/m21),2) + n22 * math.log((n22/m22),2))
        except:
            LogLikelihood=0
        return(LogLikelihood)
        
    def MI(self,n_ii, n_ix_xi_tuple, n_xx):
        '''Church Mutual Information accoding to Pazienza'''
        (n_ix, n_xi) = n_ix_xi_tuple
        self.E11=n_xi*n_ix/n_xx
        self.part=n_ii/self.E11
        self.MI=math.log(self.part,2)
        return(self.MI)
        
    def MI2(self,n_ii, n_ix_xi_tuple, n_xx):
        '''Church Mutual Information Variant accoding to Pazienza'''
        (n_ix, n_xi) = n_ix_xi_tuple
        self.E11=n_xi*n_ix/n_xx
        self.part=(n_ii/self.E11)**2
        self.MI2=math.log(self.part,2)
        return(self.MI2)
        
    def MI3(self,n_ii, n_ix_xi_tuple, n_xx):
        '''Church Mutual Information Variant accoding to Pazienza'''
        (n_ix, n_xi) = n_ix_xi_tuple
        self.E11=n_xi*n_ix/n_xx
        self.part=(n_ii/self.E11)**3
        self.MI3=math.log(self.part,2)
        return(self.MI3)
        
    def odds(self,n_ii, n_ix_xi_tuple, n_xx):
        '''Odds ratio according to NSP'''
        (n_ix, n_xi) = n_ix_xi_tuple
        n_oo=n_xx-n_ix-n_xi
        n_io=n_ix-n_ii
        n_oi=n_xi-n_ii
        n_oo=n_xx-n_ix-n_xi        
        
        n11=n_ii
        n12=n_io
        n21=n_oi
        n22=n_oo
        n1p=n11+n12
        np1=n11+n21
        n2p=n21+n22
        np2=n12+n22
        npp=n_xx

        m11 = (n1p*np1/npp)
        m12 = (n1p*np2/npp)
        m21 = (np1*n2p/npp)
        m22 = (n2p*np2/npp)
        
        if n21==0:n21=1
        if n12==0:n12=1
        ODDS_RATIO = (n11*n22)/(n21*n12)
        return(ODDS_RATIO)
        
    def z_score(self,n_ii, n_ix_xi_tuple, n_xx):
        '''z-score ratio according to NSP'''
        (n_ix, n_xi) = n_ix_xi_tuple
        n_oo=n_xx-n_ix-n_xi
        n_io=n_ix-n_ii
        n_oi=n_xi-n_ii
        n_oo=n_xx-n_ix-n_xi        
        
        n11=n_ii
        n12=n_io
        n21=n_oi
        n22=n_oo
        n1p=n11+n12
        np1=n11+n21
        n2p=n21+n22
        np2=n12+n22
        npp=n_xx

        m11 = (n1p*np1/npp)
        m12 = (n1p*np2/npp)
        m21 = (np1*n2p/npp)
        m22 = (n2p*np2/npp)
       
        zscore = (n11-m11)/(math.sqrt(m11))
        return(zscore)
   

class myTrigramAssocMeasures(nltk.collocations.TrigramAssocMeasures):
    pass
    
class myQuadgramAssocMeasures(nltk.collocations.QuadgramAssocMeasures):
    pass


 
  
    
###STUFF FROM MAP EMBEDDINGS MIKEL ARTETXE###
#cupy_utils
import numpy

try:
    import cupy
except ImportError:
    cupy = None


def supports_cupy():
    return cupy is not None


def get_cupy():
    return cupy


def get_array_module(x):
    if cupy is not None:
        return cupy.get_array_module(x)
    else:
        return numpy


def asnumpy(x):
    if cupy is not None:
        return cupy.asnumpy(x)
    else:
        return numpy.asarray(x)
#embeddings

def embeddings_read(file, threshold=0, vocabulary=None, dtype='float'):
    header = file.readline().split(' ')
    count = int(header[0]) if threshold <= 0 else min(threshold, int(header[0]))
    dim = int(header[1])
    words = []
    matrix = np.empty((count, dim), dtype=dtype) if vocabulary is None else []
    for i in range(count):
        word, vec = file.readline().split(' ', 1)
        if vocabulary is None:
            words.append(word)
            matrix[i] = np.fromstring(vec, sep=' ', dtype=dtype)
        elif word in vocabulary:
            words.append(word)
            matrix.append(np.fromstring(vec, sep=' ', dtype=dtype))
    return (words, matrix) if vocabulary is None else (words, np.array(matrix, dtype=dtype))


def embeddings_write(words, matrix, file):
    m = asnumpy(matrix)
    print('%d %d' % m.shape, file=file)
    for i in range(len(words)):
        print(words[i] + ' ' + ' '.join(['%.6g' % x for x in m[i]]), file=file)


def embeddings_length_normalize(matrix):
    xp = get_array_module(matrix)
    norms = xp.sqrt(xp.sum(matrix**2, axis=1))
    norms[norms == 0] = 1
    matrix /= norms[:, xp.newaxis]


def embeddings_mean_center(matrix):
    xp = get_array_module(matrix)
    avg = xp.mean(matrix, axis=0)
    matrix -= avg


def embeddings_length_normalize_dimensionwise(matrix):
    xp = get_array_module(matrix)
    norms = xp.sqrt(xp.sum(matrix**2, axis=0))
    norms[norms == 0] = 1
    matrix /= norms


def embeddings_mean_center_embeddingwise(matrix):
    xp = get_array_module(matrix)
    avg = xp.mean(matrix, axis=1)
    matrix -= avg[:, xp.newaxis]


def embeddings_normalize(matrix, actions):
    for action in actions:
        if action == 'unit':
            embeddings_length_normalize(matrix)
        elif action == 'center':
            embeddings_mean_center(matrix)
        elif action == 'unitdim':
            embeddings_length_normalize_dimensionwise(matrix)
        elif action == 'centeremb':
            embeddings_mean_center_embeddingwise(matrix)
            
#map_embeddings

def dropout(m, p):
    if p <= 0.0:
        return m
    else:
        xp = get_array_module(m)
        mask = xp.random.rand(*m.shape) >= p
        return m*mask


def topk_mean(m, k, inplace=False):  # TODO Assuming that axis is 1
    xp = get_array_module(m)
    n = m.shape[0]
    ans = xp.zeros(n, dtype=m.dtype)
    if k <= 0:
        return ans
    if not inplace:
        m = xp.array(m)
    ind0 = xp.arange(n)
    ind1 = xp.empty(n, dtype=int)
    minimum = m.min()
    for i in range(k):
        m.argmax(axis=1, out=ind1)
        ans += m[ind0, ind1]
        m[ind0, ind1] = minimum
    return ans / k

def supervised_mapping(src_input,trg_input,src_output,trg_output,init_dictionary,encoding="utf-8",precision="fp32",cuda=False,batch_size=1000,seed=0,unsupervised_vocab=0,src_reweight=0,trg_reweight=0,dim_reduction=0,vocabulary_cutoff=0,direction="union",csls=0,threshold=0.000001,validation=None,stochastic_initial=0.1,stochastic_multiplier=2.0,stochastic_interval=50):    
    self_learning=False
    print("SUPERVISED")
    normalize=['unit', 'center', 'unit']
    #parser.set_defaults(init_dictionary=args.supervised, normalize=['unit', 'center', 'unit'], whiten=True, src_reweight=0.5, trg_reweight=0.5, src_dewhiten='src', trg_dewhiten='trg', batch_size=1000)
    normalize=['unit', 'center', 'unit']
    whiten=True
    src_reweight=0.5
    trg_reweight=0.5
    src_dewhiten='src'
    trg_dewhiten='trg'
    batch_size=1000
    cuda=False
    identical=False
    unsupervised=False
    init_identical=False
    init_numerals=False
    init_unsupervised=False
    orthogonal=False
    unconstrained=False
    self_learning=False
    verbose=False
    if precision == 'fp16':
        dtype = 'float16'
    elif precision == 'fp32':
        dtype = 'float32'
    elif precision == 'fp64':
        dtype = 'float64'

    # Read input embeddings
    print("Read input embeddings")
    srcfile = open(src_input, encoding=encoding, errors='surrogateescape')
    trgfile = open(trg_input, encoding=encoding, errors='surrogateescape')
    src_words, x = embeddings_read(srcfile, dtype=dtype)
    trg_words, z = embeddings_read(trgfile, dtype=dtype)

    # NumPy/CuPy management
    if cuda:
        if not supports_cupy():
            print('ERROR: Install CuPy for CUDA support', file=sys.stderr)
            sys.exit(-1)
        xp = get_cupy()
        x = xp.asarray(x)
        z = xp.asarray(z)
    else:
        xp = np
    xp.random.seed(seed)

    # Build word to index map
    print("Build word to index map")
    src_word2ind = {word: i for i, word in enumerate(src_words)}
    trg_word2ind = {word: i for i, word in enumerate(trg_words)}

    # STEP 0: Normalization
    print("STEP 0: Normalization")
    embeddings_normalize(x, normalize)
    embeddings_normalize(z, normalize)

    # Build the seed dictionary
    print("Build the seed dictionary")
    src_indices = []
    trg_indices = []
    
    f = open(init_dictionary, encoding=encoding, errors='surrogateescape')
    for line in f:
        src, trg = line.split()
        try:
            src_ind = src_word2ind[src]
            trg_ind = trg_word2ind[trg]
            src_indices.append(src_ind)
            trg_indices.append(trg_ind)
        except KeyError:
            pass
            #print('WARNING: OOV dictionary entry ({0} - {1})'.format(src, trg), file=sys.stderr)

    # Read validation dictionary
    if validation is not None:
        f = open(validation, encoding=encoding, errors='surrogateescape')
        validation = collections.defaultdict(set)
        oov = set()
        vocab = set()
        for line in f:
            src, trg = line.split()
            try:
                src_ind = src_word2ind[src]
                trg_ind = trg_word2ind[trg]
                validation[src_ind].add(trg_ind)
                vocab.add(src)
            except KeyError:
                oov.add(src)
        oov -= vocab  # If one of the translation options is in the vocabulary, then the entry is not an oov
        validation_coverage = len(validation) / (len(validation) + len(oov))

    

    # Allocate memory
    print("Allocate memory")
    xw = xp.empty_like(x)
    zw = xp.empty_like(z)
    src_size = x.shape[0] if vocabulary_cutoff <= 0 else min(x.shape[0], vocabulary_cutoff)
    trg_size = z.shape[0] if vocabulary_cutoff <= 0 else min(z.shape[0], vocabulary_cutoff)
    simfwd = xp.empty((batch_size, trg_size), dtype=dtype)
    simbwd = xp.empty((batch_size, src_size), dtype=dtype)
    if validation is not None:
        simval = xp.empty((len(validation.keys()), z.shape[0]), dtype=dtype)

    best_sim_forward = xp.full(src_size, -100, dtype=dtype)
    src_indices_forward = xp.arange(src_size)
    trg_indices_forward = xp.zeros(src_size, dtype=int)
    best_sim_backward = xp.full(trg_size, -100, dtype=dtype)
    src_indices_backward = xp.zeros(trg_size, dtype=int)
    trg_indices_backward = xp.arange(trg_size)
    knn_sim_fwd = xp.zeros(src_size, dtype=dtype)
    knn_sim_bwd = xp.zeros(trg_size, dtype=dtype)

    # Training loop
    print("Training loop")
    best_objective = objective = -100.
    it = 1
    last_improvement = 0
    keep_prob = stochastic_initial
    t = time.time()
    end = not self_learning
    while True:

        # Increase the keep probability if we have not improve in stochastic_interval iterations
        if it - last_improvement > stochastic_interval:
            if keep_prob >= 1.0:
                end = True
            keep_prob = min(1.0, stochastic_multiplier*keep_prob)
            last_improvement = it

        # Update the embedding mapping
        if orthogonal or not end:  # orthogonal mapping
            u, s, vt = xp.linalg.svd(z[trg_indices].T.dot(x[src_indices]))
            w = vt.T.dot(u.T)
            x.dot(w, out=xw)
            zw[:] = z
        elif unconstrained:  # unconstrained mapping
            x_pseudoinv = xp.linalg.inv(x[src_indices].T.dot(x[src_indices])).dot(x[src_indices].T)
            w = x_pseudoinv.dot(z[trg_indices])
            x.dot(w, out=xw)
            zw[:] = z
        else:  # advanced mapping

            # TODO xw.dot(wx2, out=xw) and alike not working
            xw[:] = x
            zw[:] = z

            # STEP 1: Whitening
            def whitening_transformation(m):
                u, s, vt = xp.linalg.svd(m, full_matrices=False)
                return vt.T.dot(xp.diag(1/s)).dot(vt)
            if whiten:
                wx1 = whitening_transformation(xw[src_indices])
                wz1 = whitening_transformation(zw[trg_indices])
                xw = xw.dot(wx1)
                zw = zw.dot(wz1)

            # STEP 2: Orthogonal mapping
            wx2, s, wz2_t = xp.linalg.svd(xw[src_indices].T.dot(zw[trg_indices]))
            wz2 = wz2_t.T
            xw = xw.dot(wx2)
            zw = zw.dot(wz2)

            # STEP 3: Re-weighting
            xw *= s**src_reweight
            zw *= s**trg_reweight

            # STEP 4: De-whitening
            if src_dewhiten == 'src':
                xw = xw.dot(wx2.T.dot(xp.linalg.inv(wx1)).dot(wx2))
            elif src_dewhiten == 'trg':
                xw = xw.dot(wz2.T.dot(xp.linalg.inv(wz1)).dot(wz2))
            if trg_dewhiten == 'src':
                zw = zw.dot(wx2.T.dot(xp.linalg.inv(wx1)).dot(wx2))
            elif trg_dewhiten == 'trg':
                zw = zw.dot(wz2.T.dot(xp.linalg.inv(wz1)).dot(wz2))

            # STEP 5: Dimensionality reduction
            if dim_reduction > 0:
                xw = xw[:, :dim_reduction]
                zw = zw[:, :dim_reduction]

        # Self-learning
        if end:
            break
        else:
            # Update the training dictionary
            if direction in ('forward', 'union'):
                if csls_neighborhood > 0:
                    for i in range(0, trg_size, simbwd.shape[0]):
                        j = min(i + simbwd.shape[0], trg_size)
                        zw[i:j].dot(xw[:src_size].T, out=simbwd[:j-i])
                        knn_sim_bwd[i:j] = topk_mean(simbwd[:j-i], k=csls_neighborhood, inplace=True)
                for i in range(0, src_size, simfwd.shape[0]):
                    j = min(i + simfwd.shape[0], src_size)
                    xw[i:j].dot(zw[:trg_size].T, out=simfwd[:j-i])
                    simfwd[:j-i].max(axis=1, out=best_sim_forward[i:j])
                    simfwd[:j-i] -= knn_sim_bwd/2  # Equivalent to the real CSLS scores for NN
                    dropout(simfwd[:j-i], 1 - keep_prob).argmax(axis=1, out=trg_indices_forward[i:j])
            if direction in ('backward', 'union'):
                if csls_neighborhood > 0:
                    for i in range(0, src_size, simfwd.shape[0]):
                        j = min(i + simfwd.shape[0], src_size)
                        xw[i:j].dot(zw[:trg_size].T, out=simfwd[:j-i])
                        knn_sim_fwd[i:j] = topk_mean(simfwd[:j-i], k=csls_neighborhood, inplace=True)
                for i in range(0, trg_size, simbwd.shape[0]):
                    j = min(i + simbwd.shape[0], trg_size)
                    zw[i:j].dot(xw[:src_size].T, out=simbwd[:j-i])
                    simbwd[:j-i].max(axis=1, out=best_sim_backward[i:j])
                    simbwd[:j-i] -= knn_sim_fwd/2  # Equivalent to the real CSLS scores for NN
                    dropout(simbwd[:j-i], 1 - keep_prob).argmax(axis=1, out=src_indices_backward[i:j])
            if direction == 'forward':
                src_indices = src_indices_forward
                trg_indices = trg_indices_forward
            elif direction == 'backward':
                src_indices = src_indices_backward
                trg_indices = trg_indices_backward
            elif direction == 'union':
                src_indices = xp.concatenate((src_indices_forward, src_indices_backward))
                trg_indices = xp.concatenate((trg_indices_forward, trg_indices_backward))

            # Objective function evaluation
            if direction == 'forward':
                objective = xp.mean(best_sim_forward).tolist()
            elif direction == 'backward':
                objective = xp.mean(best_sim_backward).tolist()
            elif direction == 'union':
                objective = (xp.mean(best_sim_forward) + xp.mean(best_sim_backward)).tolist() / 2
            if objective - best_objective >= threshold:
                last_improvement = it
                best_objective = objective

            # Accuracy and similarity evaluation in validation
            if validation is not None:
                src = list(validation.keys())
                xw[src].dot(zw.T, out=simval)
                nn = asnumpy(simval.argmax(axis=1))
                accuracy = np.mean([1 if nn[i] in validation[src[i]] else 0 for i in range(len(src))])
                similarity = np.mean([max([simval[i, j].tolist() for j in validation[src[i]]]) for i in range(len(src))])

            # Logging
            duration = time.time() - t
            if verbose:
                print(file=sys.stderr)
                print('ITERATION {0} ({1:.2f}s)'.format(it, duration), file=sys.stderr)
                print('\t- Objective:        {0:9.4f}%'.format(100 * objective), file=sys.stderr)
                print('\t- Drop probability: {0:9.4f}%'.format(100 - 100*keep_prob), file=sys.stderr)
                if validation is not None:
                    print('\t- Val. similarity:  {0:9.4f}%'.format(100 * similarity), file=sys.stderr)
                    print('\t- Val. accuracy:    {0:9.4f}%'.format(100 * accuracy), file=sys.stderr)
                    print('\t- Val. coverage:    {0:9.4f}%'.format(100 * validation_coverage), file=sys.stderr)
                sys.stderr.flush()
            if log is not None:
                val = '{0:.6f}\t{1:.6f}\t{2:.6f}'.format(
                    100 * similarity, 100 * accuracy, 100 * validation_coverage) if validation is not None else ''
                print('{0}\t{1:.6f}\t{2}\t{3:.6f}'.format(it, 100 * objective, val, duration), file=log)
                log.flush()

        t = time.time()
        it += 1

    # Write mapped embeddings
    print("Write mapped embeddings")
    srcfile = open(src_output, mode='w', encoding=encoding, errors='surrogateescape')
    trgfile = open(trg_output, mode='w', encoding=encoding, errors='surrogateescape')
    embeddings_write(src_words, xw, srcfile)
    embeddings_write(trg_words, zw, trgfile)
    srcfile.close()
    trgfile.close() 

